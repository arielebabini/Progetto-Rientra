"""
rdf2rdb_no_fine_tuned.py
========================
Step 2 del progetto Rientra@: RDB2RDF con matching verso l'ontologia principale.

Dataset in ingresso:
  - Tabella PostgreSQL  : ext_job  (ext01…ext09)
  - Ontologia           : Rientra.rdf  (namespace http://www.stiima.cnr.it/JobList#)

Strategie di matching:
  S1 — String Matching          : Jaccard + Overlap + Levenshtein
                                   Pre-processing esteso: normalizzazione simboli,
                                   espansione abbreviazioni, rimozione stopwords.
  S2 — all-mpnet-base-v2        : sentence-transformer nativo (768 dim), cosine similarity.
                                   Superiore a MiniLM su testi lunghi. Nessun problema
                                   di anisotropy perché ottimizzato per sentence similarity.
  S4 — all-MiniLM-L6-v2        : sentence-transformer compatto (384 dim), riferimento.

Nota sui modelli:
  Tutti i modelli NLP usati (S2, S4) sono sentence-transformers nativi,
  non encoder BERT raw. Questo garantisce embedding di frase calibrati per
  cosine similarity e assenza del fenomeno di anisotropy.

Requisiti:
    pip install psycopg2-binary rdflib rich sentence-transformers openpyxl

Uso:
    python rdf2rdb_no_fine_tuned.py
"""

import os
import sys
import numpy as np
from datetime import datetime

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich import box
from rich.text import Text
from rich.rule import Rule

console = Console()

# ─────────────────────────────────────────────────────────────
# CONFIGURAZIONE
# ─────────────────────────────────────────────────────────────

DB_CONFIG = {
    "host":     "localhost",
    "port":     5432,
    "dbname":   "rientra_db",
    "user":     "postgres",
    "password": "postgres",
}

ONTOLOGY_FILE    = "Rientra.rdf"
OUTPUT_DIR       = "output"

JOBLIST_BASE = "http://www.stiima.cnr.it/JobList#"
RIENTRA_BASE = "https://www.stiima.cnr.it/rientra#"

# Soglie cosine similarity
# Tutti i modelli NLP sono sentence-transformers → stessa scala di score
S1_THRESHOLD        = 0.12
MPNET_THRESHOLD     = 0.50   # all-mpnet-base-v2: sentence-transformer nativo
MINILM_THRESHOLD    = 0.50   # all-MiniLM-L6-v2: riferimento

# Modelli — tutti sentence-transformers nativi
MODELS = {
    "S2_MPNET":  "sentence-transformers/all-mpnet-base-v2",
    "S4_MiniLM": "all-MiniLM-L6-v2",
}

# ─────────────────────────────────────────────────────────────
# AI JUDGE — modello Ollama locale
# Modelli consigliati per MacBook Pro M1 16 GB:
#   gemma3:4b    ~3 GB RAM  — veloce, buona qualità (DEFAULT)
#   qwen2.5:7b   ~5 GB RAM  — top reasoning
#   llama3.1:8b  ~6 GB RAM  — bilanciato
# Per usare un altro modello: cambia OLLAMA_MODEL qui sotto.
# ─────────────────────────────────────────────────────────────
OLLAMA_MODEL = "mistral:latest"

# Abbreviazioni comuni nel dominio IT/lavoro da espandere nel pre-processing S1
ABBREVIATIONS = {
    "ml":   "machine learning",
    "ai":   "artificial intelligence",
    "nlp":  "natural language processing",
    "hr":   "human resources",
    "pr":   "public relations",
    "it":   "information technology",
    "db":   "database",
    "mgr":  "manager",
    "mgmt": "management",
    "admin": "administration",
    "asst": "assistant",
    "assoc": "associate",
    "spec": "specialist",
    "coord": "coordinator",
    "dept": "department",
    "exec": "executive",
    "sr":   "senior",
    "jr":   "junior",
    "r&d":  "research and development",
    "b2b":  "business to business",
    "b2c":  "business to consumer",
    "erp":  "enterprise resource planning",
    "crm":  "customer relationship management",
    "ui":   "user interface",
    "ux":   "user experience",
    "qa":   "quality assurance",
    "ops":  "operations",
    "dev":  "developer",
    "sw":   "software",
    "hw":   "hardware",
}


# ─────────────────────────────────────────────────────────────
# HELPERS VISIVI
# ─────────────────────────────────────────────────────────────

def score_color(score: float) -> str:
    if score >= 0.85: return "bold green"
    if score >= 0.70: return "green"
    if score >= 0.50: return "yellow"
    if score >= 0.30: return "orange3"
    return "red"

def score_bar(score: float, width: int = 8) -> str:
    filled = round(score * width)
    return "█" * filled + "░" * (width - filled)


# ═════════════════════════════════════════════════════════════
# UTILITÀ DB
# ═════════════════════════════════════════════════════════════

def get_connection():
    return psycopg2.connect(**DB_CONFIG)

def fetch_all(sql: str, params=None) -> list[dict]:
    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            return [dict(r) for r in cur.fetchall()]

def test_connection() -> bool:
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT version();")
                v = cur.fetchone()[0]
        console.print(f"  [bold green]✓[/] PostgreSQL [dim]{v[:60]}...[/]")
        return True
    except Exception as e:
        console.print(f"  [bold red]✗[/] Errore connessione: [red]{e}[/]")
        return False


# ═════════════════════════════════════════════════════════════
# ESTRAZIONE VOCABOLARIO DALL'ONTOLOGIA
# ═════════════════════════════════════════════════════════════

def load_ontology(path: str):
    """Carica il file RDF e restituisce un oggetto Graph rdflib."""
    from rdflib import Graph
    g = Graph()
    g.parse(path, format="xml")
    return g

def extract_ontology_jobs(onto_graph) -> list[dict]:
    from rdflib import RDF, OWL
    from rdflib.namespace import RDFS
    jobs = []
    for s, _, _ in onto_graph.triples((None, RDF.type, OWL.Class)):
        if not str(s).startswith(JOBLIST_BASE):
            continue
        local = str(s).split("#")[1]
        if local in ("Job", "Job_Descriptor"):
            continue

        label = str(onto_graph.value(s, RDFS.label) or "").strip()
        titles, desc = "", ""
        for pred, obj in onto_graph.predicate_objects(s):
            ps = str(pred)
            if "Net_job_titles" in ps or "job_titles" in ps:
                titles = str(obj)
            if "Net_short_description" in ps or "short_description" in ps:
                desc = str(obj)

        if titles and ":" in titles:
            titles = titles.split(":", 1)[1].strip()

        display = label or local.replace("_", " ")
        jobs.append({
            "uri":         str(s),
            "local_name":  local,
            "label":       display,
            "titles":      titles,
            "description": desc,
        })

    jobs.sort(key=lambda x: x["label"])
    return jobs

def _build_onto_text(j: dict) -> str:
    """Testo composito per l'ontologia: label + titoli alternativi + descrizione."""
    parts = [j["label"]]
    if j["titles"]:
        parts.append(j["titles"])
    if j["description"]:
        parts.append(j["description"])
    return " ".join(parts)


# ═════════════════════════════════════════════════════════════
# STRATEGIA 1 — STRING MATCHING (pre-processing esteso)
# ═════════════════════════════════════════════════════════════

# Stopwords leggere per il dominio job matching
# (parole molto frequenti che non aggiungono discriminazione)
_STOPWORDS = {
    "and", "or", "the", "a", "an", "of", "in", "for", "to",
    "with", "at", "by", "as", "on", "is", "are", "be", "been",
    "from", "that", "this", "which", "who", "other", "related",
}

def _norm(s: str) -> str:
    """
    Pre-processing esteso per string matching.

    Passi applicati in ordine:
      1. Lowercase e strip degli spazi
      2. Sostituzione dei separatori (-, /, _, .) con spazio
      3. Rimozione della punteggiatura residua (',', '(', ')', ecc.)
      4. Espansione delle abbreviazioni comuni (ML → machine learning)
      5. Rimozione delle stopwords leggere
      6. Collasso degli spazi multipli
    """
    import re
    s = s.lower().strip()
    # Separatori → spazio
    s = re.sub(r"[-/_.]", " ", s)
    # Punteggiatura residua
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    # Espansione abbreviazioni (parola intera, non sottostringa)
    tokens = s.split()
    expanded = []
    for tok in tokens:
        expanded.append(ABBREVIATIONS.get(tok, tok))
    # Rimozione stopwords + collasso spazi
    tokens = " ".join(expanded).split()
    tokens = [t for t in tokens if t not in _STOPWORDS and len(t) > 1]
    return " ".join(tokens)

def _jaccard(a: str, b: str) -> float:
    a, b = set(_norm(a).split()), set(_norm(b).split())
    return len(a & b) / len(a | b) if (a | b) else 1.0

def _overlap(a: str, b: str) -> float:
    a, b = set(_norm(a).split()), set(_norm(b).split())
    return len(a & b) / min(len(a), len(b)) if (a and b) else 0.0

def _levenshtein(s1: str, s2: str) -> float:
    s1, s2 = _norm(s1), _norm(s2)
    m, n = len(s1), len(s2)
    if not m and not n:
        return 1.0
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev, dp[0] = dp[:], i
        for j in range(1, n + 1):
            cost = 0 if s1[i-1] == s2[j-1] else 1
            dp[j] = min(prev[j] + 1, dp[j-1] + 1, prev[j-1] + cost)
    return 1.0 - dp[n] / max(m, n)

def _combined(a: str, b: str) -> float:
    return 0.5 * _jaccard(a, b) + 0.3 * _overlap(a, b) + 0.2 * _levenshtein(a, b)

def s1_match_one(title: str, onto_jobs: list[dict]) -> dict | None:
    best, bj, bv = 0.0, None, None
    for j in onto_jobs:
        cands = [j["label"]]
        if j["titles"]:
            cands += [t.strip() for t in j["titles"].split(",") if t.strip()]
        for c in cands:
            sc = _combined(title, c)
            if sc > best:
                best, bj, bv = sc, j, c
    if best >= S1_THRESHOLD:
        return {"job": bj, "score": best, "via": bv}
    return None

def run_string_matching(db_jobs: list[dict], onto_jobs: list[dict]) -> list[dict]:
    results = []
    with Progress(SpinnerColumn(), TextColumn("{task.description}"),
                  BarColumn(), TaskProgressColumn(),
                  console=console, transient=True) as p:
        task = p.add_task("  Calcolo...", total=len(db_jobs))
        for r in db_jobs:
            results.append({"id": r["id"], "title": r["title"],
                            "match_s1": s1_match_one(r["title"], onto_jobs)})
            p.advance(task)
    return results


# ═════════════════════════════════════════════════════════════
# COSINE SIMILARITY — funzione centrale condivisa da S2 e S3
# ═════════════════════════════════════════════════════════════

def cosine_similarity_matrix(query_emb: np.ndarray,
                              corpus_embs: np.ndarray) -> np.ndarray:
    """
    Calcola la cosine similarity tra un vettore query e una matrice di vettori corpus.

    Formula:
        cos(θ) = (A · B) / (‖A‖ · ‖B‖)

    dove A è il vettore del job nel DB e B è il vettore di una professione O*NET.

    Il risultato è compreso tra -1 e 1:
      1.0  → testi identici nel significato
      0.0  → testi ortogonali (nessuna relazione semantica)
     -1.0  → testi opposti (raro con embedding linguistici)

    Per embedding BERT non negativi il range effettivo è [0, 1].
    """
    query_norm  = query_emb / (np.linalg.norm(query_emb) + 1e-9)
    corpus_norm = corpus_embs / (np.linalg.norm(corpus_embs, axis=1, keepdims=True) + 1e-9)
    return corpus_norm.dot(query_norm)


# ═════════════════════════════════════════════════════════════
# EMBEDDING CON BERT RAW (BERT base e BioBERT)
# ═════════════════════════════════════════════════════════════

# ═════════════════════════════════════════════════════════════
# SENTENCE-TRANSFORMER MATCHING — funzione generica per S2, S3, S4
# ═════════════════════════════════════════════════════════════

def run_st_matching(db_jobs: list[dict],
                    onto_jobs: list[dict],
                    model_name_or_path: str,
                    strategy_key: str,
                    threshold: float) -> tuple[list[dict], str]:
    """
    Esegue il matching con un qualsiasi modello sentence-transformers.

    Usato per S2 (all-mpnet-base-v2), S3 (MPNet fine-tuned) e S4 (MiniLM).
    Tutti questi modelli sono sentence-transformers nativi: il loro .encode()
    produce embedding ottimizzati per cosine similarity senza bisogno di
    estrarre manualmente il token [CLS] o applicare pooling.

    La cosine similarity è calcolata esplicitamente con cosine_similarity_matrix
    per uniformità metodologica tra tutte le strategie NLP.
    """
    try:
        from sentence_transformers import SentenceTransformer

        label = model_name_or_path.split("/")[-1]
        with console.status(
                f"  Caricamento [cyan]{label}[/]...", spinner="dots"):
            model = SentenceTransformer(model_name_or_path)
        console.print(
            f"  [bold green]✓[/] Modello caricato: [cyan]{model_name_or_path}[/]")

        onto_texts = [_build_onto_text(j) for j in onto_jobs]
        db_texts   = [f"{r['title']}. {r.get('description', '')}"
                      for r in db_jobs]

        with Progress(SpinnerColumn(), TextColumn("{task.description}"),
                      BarColumn(), TaskProgressColumn(),
                      console=console, transient=True) as p:
            t1 = p.add_task("  Encoding O*NET...", total=1)
            onto_embs = model.encode(onto_texts, convert_to_numpy=True)
            p.advance(t1)

            results = []
            t2 = p.add_task("  Cosine similarity...", total=len(db_jobs))
            for i, row in enumerate(db_jobs):
                q_emb = model.encode(db_texts[i], convert_to_numpy=True)
                sims  = cosine_similarity_matrix(q_emb, onto_embs)
                idx   = int(np.argmax(sims))
                sc    = float(sims[idx])
                m = {"job": onto_jobs[idx], "score": sc} \
                    if sc >= threshold else None
                results.append({
                    "id":        row["id"],
                    "title":     row["title"],
                    "match":     m,
                    "score_raw": sc,
                    "strategy":  strategy_key,
                })
                p.advance(t2)

        return results, model_name_or_path

    except Exception as e:
        console.print(
            f"  [yellow]⚠[/] {model_name_or_path} non disponibile: [dim]{e}[/]")
        return None, None




# ═════════════════════════════════════════════════════════════
# EMBEDDING CON SENTENCE-TRANSFORMERS (MiniLM — S4)
# ═════════════════════════════════════════════════════════════

def run_minilm_matching(db_jobs: list[dict],
                        onto_jobs: list[dict],
                        threshold: float) -> tuple[list[dict], str]:
    """
    Matching con all-MiniLM-L6-v2 (modello di riferimento compatto).
    Delega a run_st_matching con strategy_key S4_MiniLM.
    """
    return run_st_matching(
        db_jobs, onto_jobs,
        MODELS["S4_MiniLM"], "S4_MiniLM", threshold)


# ═════════════════════════════════════════════════════════════
# OUTPUT RICH
# ═════════════════════════════════════════════════════════════

def print_ontology_table(onto_jobs: list[dict]):
    t = Table(title="Professioni O*NET nell'ontologia",
              box=box.ROUNDED, header_style="bold cyan",
              title_style="bold white", show_lines=False)
    t.add_column("#",           style="dim",       width=3,  justify="right")
    t.add_column("URI locale",  style="cyan",       width=46)
    t.add_column("Label O*NET", style="bold white", width=38)
    t.add_column("Desc.",       justify="center",   width=6)
    for i, j in enumerate(onto_jobs, 1):
        t.add_row(str(i), j["local_name"], j["label"],
                  "[green]✓[/]" if j["description"] else "[red]✗[/]")
    console.print(t)


def print_db_table(db_jobs: list[dict]):
    t = Table(title="Lavori nel DB esterno (ext_job)",
              box=box.ROUNDED, header_style="bold magenta",
              title_style="bold white", show_lines=False)
    t.add_column("ID",    style="bold magenta", width=6)
    t.add_column("Titolo",                      width=30)
    t.add_column("Descrizione (estratto)",      width=62)
    for r in db_jobs:
        desc = (r.get("description") or "")[:60] + "…"
        t.add_row(r["id"], r["title"], f"[dim]{desc}[/]")
    console.print(t)


def _score_cell(score: float | None, raw: float = 0.0) -> tuple[Text, Text]:
    """Restituisce (cella_label_aggiuntiva, cella_score) per la tabella."""
    if score is not None:
        bar = Text()
        bar.append(f"{score:.3f}\n", style=score_color(score))
        bar.append(score_bar(score),  style=score_color(score))
        return Text(""), bar
    else:
        lbl = Text(f"(raw: {raw:.3f})", style="dim red")
        bar = Text()
        bar.append(f"{raw:.3f}\n", style="dim red")
        bar.append(score_bar(raw),  style="dim red")
        return lbl, bar


def print_full_comparison(s1_res: list[dict],
                          nlp_results: dict[str, list[dict]],
                          model_names: dict[str, str]):
    """
    Tabella unica con S1 + tutti i modelli NLP disponibili, una colonna per modello.
    nlp_results: { "S2_MPNET": [...], "S3_FINETUNED": [...], "S4_MiniLM": [...] }
    model_names: { "S2_MPNET": "all-mpnet-base-v2", ... }
    """
    # Indici per lookup rapido per ogni strategia
    idx = {key: {r["id"]: r for r in res}
           for key, res in nlp_results.items() if res}

    available_strategies = [k for k in ["S2_MPNET", "S4_MiniLM"]
                            if k in idx]

    # Costruisce header dinamico
    title_parts = ["S1 String"] + [
        model_names.get(k, k).split("/")[-1][:18]
        for k in available_strategies
    ]
    t = Table(
        title="Confronto strategie di matching · cosine similarity",
        box=box.SIMPLE_HEAD,
        show_lines=True,
        header_style="bold white on dark_blue",
        title_style="bold white",
        expand=True,
    )
    t.add_column("ID",       style="bold", width=6,  justify="center")
    t.add_column("Titolo DB",              width=26)
    t.add_column("S1 — Match",             width=28)
    t.add_column("Score S1", justify="center", width=14)

    for k in available_strategies:
        short = model_names.get(k, k).split("/")[-1][:20]
        t.add_column(f"{k[:2]} — {short}", width=28)
        t.add_column(f"Score {k[:2]}", justify="center", width=14)

    t.add_column("Verdetto", justify="center", width=10)

    for r in s1_res:
        rid = r["id"]
        m1  = r.get("match_s1")

        # S1
        if m1:
            s1_cell = Text()
            s1_cell.append(m1["job"]["label"] + "\n")
            s1_cell.append(f"via: {m1['via'][:24]}", style="dim")
            s1_sc = Text()
            s1_sc.append(f"{m1['score']:.3f}\n", style=score_color(m1["score"]))
            s1_sc.append(score_bar(m1["score"]),  style=score_color(m1["score"]))
        else:
            s1_cell = Text("— nessun match —", style="dim")
            s1_sc   = Text("—", style="dim")

        row_cells = [rid, r["title"], s1_cell, s1_sc]

        # NLP strategies
        nlp_matches = []
        for k in available_strategies:
            r2   = idx[k].get(rid, {})
            m2   = r2.get("match")
            raw  = r2.get("score_raw", 0.0)
            if m2:
                cell = Text(m2["job"]["label"])
                sc   = Text()
                sc.append(f"{m2['score']:.3f}\n", style=score_color(m2["score"]))
                sc.append(score_bar(m2["score"]),  style=score_color(m2["score"]))
                nlp_matches.append(m2["job"]["uri"])
            else:
                cell = Text()
                cell.append("— sotto soglia —\n", style="dim")
                cell.append(f"(raw: {raw:.3f})", style="dim red")
                sc   = Text()
                sc.append(f"{raw:.3f}\n", style="dim red")
                sc.append(score_bar(raw),  style="dim red")
                nlp_matches.append(None)
            row_cells += [cell, sc]

        # Verdetto: consensus tra i modelli che hanno trovato un match
        valid = [u for u in nlp_matches if u]
        if not valid and not m1:
            verdetto = Text("✗ no", style="bold red")
        elif len(set(valid)) == 1 and len(valid) == len(available_strategies):
            verdetto = Text("✓ tutti", style="bold green")
        elif len(set(valid)) == 1 and valid:
            verdetto = Text("~ parz.", style="yellow")
        elif valid:
            verdetto = Text("≠ div.", style="orange3")
        else:
            verdetto = Text("S1 only", style="magenta")

        row_cells.append(verdetto)
        t.add_row(*row_cells)

    console.print(t)
    console.print(
        "  [bold green]✓ tutti[/] tutti i modelli NLP concordano  "
        "[yellow]~ parz.[/] accordo parziale  "
        "[orange3]≠ div.[/] modelli divergono  "
        "[magenta]S1 only[/] solo string match  "
        "[bold red]✗ no[/] nessun match\n"
    )


def print_cosine_detail(nlp_results: dict[str, list[dict]],
                        model_names: dict[str, str],
                        onto_jobs: list[dict]):
    """
    Tabella dettaglio cosine similarity: per ogni job DB mostra lo score
    con TUTTE le professioni O*NET, per ogni modello disponibile.
    Utile per capire la distribuzione degli score e validare le soglie.
    """
    console.print(Rule("[dim]Dettaglio score cosine per tutti i candidati O*NET[/]", style="dim"))
    console.print("[dim]  (i valori in verde sono quelli scelti come match migliore)[/]\n")

    available = [k for k in ["S2_MPNET", "S4_MiniLM"]
                 if k in nlp_results and nlp_results[k]]

    for strategy_key in available:
        res_list = nlp_results[strategy_key]
        mname    = model_names.get(strategy_key, strategy_key)
        console.print(f"  [bold cyan]{strategy_key}[/] · [dim]{mname}[/]")

        t = Table(box=box.MINIMAL, show_header=True,
                  header_style="dim", show_lines=False)
        t.add_column("Professione O*NET", width=38)
        for r in res_list:
            t.add_column(r["title"][:14], justify="right", width=10)

        # Nota: il dettaglio per O*NET richiede gli score completi
        # che non sono salvati nella struttura corrente.
        # Mostriamo solo il best match per job.
        idx_map = {r["id"]: r for r in res_list}
        for j in onto_jobs:
            row_vals = [j["label"][:37]]
            for r in res_list:
                m   = r.get("match")
                raw = r.get("score_raw", 0.0)
                if m and m["job"]["uri"] == j["uri"]:
                    row_vals.append(f"[bold green]{m['score']:.3f}[/]")
                elif not m and r.get("score_raw", 0) > 0:
                    # Se questo è il candidato con score più alto (anche sotto soglia)
                    row_vals.append(f"[dim]{raw:.3f}[/]")
                else:
                    row_vals.append("[dim]—[/]")
            t.add_row(*row_vals)

        console.print(t)
        console.print()


def print_summary_multi(s1_res: list[dict],
                        nlp_results: dict[str, list[dict]],
                        model_names: dict[str, str]):
    n_total = len(s1_res)
    n_s1    = sum(1 for r in s1_res if r.get("match_s1"))

    lines = [f"  Job totali nel DB  : [bold]{n_total}[/]",
             f"  Match da S1        : [bold {'green' if n_s1 else 'red'}]{n_s1}/{n_total}[/]  [dim](soglia {S1_THRESHOLD})[/]"]

    for key in ["S2_MPNET", "S4_MiniLM"]:
        res = nlp_results.get(key)
        if not res:
            lines.append(f"  Match da {key[:2]}         : [dim]— modello non disponibile —[/]")
            continue
        _thresholds = {
            "S2_MPNET":  MPNET_THRESHOLD,
            "S4_MiniLM": MINILM_THRESHOLD,
        }
        thr = _thresholds.get(key, 0.50)
        n   = sum(1 for r in res if r.get("match"))
        mname = model_names.get(key, key).split("/")[-1]
        lines.append(
            f"  Match da {key[:2]}         : [bold {'green' if n else 'red'}]{n}/{n_total}[/]"
            f"  [dim](soglia {thr} · {mname})[/]"
        )

    lines += [
        "",
        "  [dim]Modelli NLP: S2=all-mpnet-base-v2 · S4=all-MiniLM-L6-v2[/]",
        "  [dim]Tutti sentence-transformers nativi[/]",
    ]

    console.print(Panel("\n".join(lines),
                        title="[bold white]Riepilogo",
                        border_style="cyan", padding=(1, 2)))


# ═════════════════════════════════════════════════════════════
# OUTPUT EXCEL
# ═════════════════════════════════════════════════════════════

# ── Stili riutilizzabili ──────────────────────────────────────
_HDR_FILL   = PatternFill("solid", fgColor="1E3264")   # blu scuro
_HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_BODY_FONT  = Font(name="Arial", size=10)
_WRAP       = Alignment(wrap_text=True, vertical="top")
_CENTER     = Alignment(horizontal="center", vertical="center")
_THIN       = Side(style="thin", color="CCCCCC")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FILL_S1    = PatternFill("solid", fgColor="FFF3CD")   # giallo tenue
_FILL_BERT  = PatternFill("solid", fgColor="D0E8FB")   # azzurro tenue
_FILL_BIO   = PatternFill("solid", fgColor="F3D0FB")   # lilla tenue
_FILL_MINI  = PatternFill("solid", fgColor="D0FBE0")   # verde tenue
_FILL_ALT   = PatternFill("solid", fgColor="F8F8F8")   # grigio alternato

_GREEN_FONT = Font(name="Arial", size=10, color="1D6A38", bold=True)
_RED_FONT   = Font(name="Arial", size=10, color="A32D2D")
_DIM_FONT   = Font(name="Arial", size=10, color="888888", italic=True)


def _set_header(ws, row: int, cols: list[tuple[int, str]]):
    """Scrive una riga di intestazione con stile."""
    for col, text in cols:
        c = ws.cell(row=row, column=col, value=text)
        c.font      = _HDR_FONT
        c.fill      = _HDR_FILL
        c.alignment = _CENTER
        c.border    = _BORDER


def _style_cell(cell, fill=None, font=None, wrap=False, align=None):
    if fill:  cell.fill      = fill
    if font:  cell.font      = font or _BODY_FONT
    if not font: cell.font   = _BODY_FONT
    cell.border    = _BORDER
    cell.alignment = align or (_WRAP if wrap else Alignment(vertical="top"))


def save_excel(db_jobs: list[dict],
               s1_res:  list[dict],
               nlp_results: dict[str, list[dict]],
               model_names: dict[str, str],
               onto_jobs: list[dict],
               judge_results: list[dict] | None = None) -> str:
    """
    Produce un file Excel con quattro fogli:
      1. Matching Results  — una riga per job, colonne per ogni strategia
      2. Ontology Jobs     — le 13 professioni O*NET con dettagli
      3. Score Detail      — tutti gli score raw per ogni job × strategia
      4. AI Judge          — verdetti del judge LLM locale (se disponibili)
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = os.path.join(OUTPUT_DIR, "rientra_matching.xlsx")
    wb   = Workbook()

    # Indici per lookup veloce
    s1_idx  = {r["id"]: r for r in s1_res}
    nlp_idx = {key: {r["id"]: r for r in res}
               for key, res in nlp_results.items() if res}
    avail   = [k for k in ["S2_MPNET", "S4_MiniLM"] if k in nlp_idx]

    # ── FOGLIO 1: Matching Results ────────────────────────────
    ws1 = wb.active
    ws1.title = "Matching Results"
    ws1.freeze_panes = "A2"

    # Costruisce header dinamico
    base_cols = [
        (1, "ID"),
        (2, "Job Title (DB)"),
        (3, "Description (DB)"),
        (4, "IRI individuo"),
        # S1
        (5,  "S1 · Match O*NET"),
        (6,  "S1 · Score"),
        (7,  "S1 · Via label"),
        (8,  "S1 · O*NET URI"),
    ]
    nlp_start = 9
    nlp_col_map: dict[str, dict] = {}  # key → {match, score, uri}
    col = nlp_start
    for key in avail:
        short = model_names.get(key, key).split("/")[-1]
        nlp_col_map[key] = {"match": col, "score": col+1, "uri": col+2}
        base_cols += [
            (col,   f"{key[:2]} · {short} Match"),
            (col+1, f"{key[:2]} · Score"),
            (col+2, f"{key[:2]} · O*NET URI"),
        ]
        col += 3

    verdict_col = col
    base_cols.append((verdict_col, "Verdetto"))

    _set_header(ws1, 1, base_cols)

    # Larghezze colonne foglio 1
    col_widths_1 = {1:8, 2:28, 3:55, 4:35, 5:32, 6:10, 7:28, 8:55}
    for key in avail:
        cm = nlp_col_map[key]
        col_widths_1[cm["match"]] = 32
        col_widths_1[cm["score"]] = 10
        col_widths_1[cm["uri"]]   = 55
    col_widths_1[verdict_col] = 12
    for c, w in col_widths_1.items():
        ws1.column_dimensions[get_column_letter(c)].width = w

    for r_idx, row in enumerate(db_jobs, start=2):
        rid  = row["id"]
        alt  = (r_idx % 2 == 0)
        fill = _FILL_ALT if alt else None

        def wc(col, val, cfill=None, cfont=None, wrap=True):
            c = ws1.cell(row=r_idx, column=col, value=val)
            _style_cell(c, fill=cfill or fill, font=cfont, wrap=wrap)

        wc(1, rid,           wrap=False)
        wc(2, row["title"])
        wc(3, row.get("description",""))
        wc(4, f"{RIENTRA_BASE}ExtJob_{rid}")

        # S1
        m1 = s1_idx.get(rid, {}).get("match_s1")
        if m1:
            wc(5, m1["job"]["label"],   cfill=_FILL_S1)
            wc(6, round(m1["score"],4), cfill=_FILL_S1, wrap=False)
            wc(7, m1["via"],            cfill=_FILL_S1)
            wc(8, m1["job"]["uri"],     cfill=_FILL_S1)
        else:
            for c in [5,6,7,8]:
                cell = ws1.cell(row=r_idx, column=c, value="—")
                _style_cell(cell, fill=fill, font=_DIM_FONT, wrap=False)

        # NLP strategies
        uri_set = set()
        for key in avail:
            cm    = nlp_col_map[key]
            r2    = nlp_idx[key].get(rid, {})
            m2    = r2.get("match")
            raw   = r2.get("score_raw", 0.0)
            fills = {"S2_MPNET":  _FILL_BERT,
                     "S4_MiniLM": _FILL_MINI}
            sf = fills.get(key)
            if m2:
                wc(cm["match"], m2["job"]["label"],    cfill=sf)
                sc_cell = ws1.cell(row=r_idx, column=cm["score"],
                                   value=round(m2["score"],4))
                _style_cell(sc_cell, fill=sf, wrap=False)
                sc_cell.font = _GREEN_FONT if m2["score"] >= 0.70 else _BODY_FONT
                wc(cm["uri"],   m2["job"]["uri"],       cfill=sf)
                uri_set.add(m2["job"]["uri"])
            else:
                no_cell = ws1.cell(row=r_idx, column=cm["match"],
                                   value=f"sotto soglia (raw {raw:.3f})")
                _style_cell(no_cell, fill=fill, font=_DIM_FONT)
                ws1.cell(row=r_idx, column=cm["score"],
                         value=round(raw,4)).font = _RED_FONT
                _style_cell(ws1.cell(row=r_idx, column=cm["score"]),
                            fill=fill, wrap=False)
                ws1.cell(row=r_idx, column=cm["uri"], value="—")
                _style_cell(ws1.cell(row=r_idx, column=cm["uri"]),
                            fill=fill, font=_DIM_FONT)

        # Verdetto
        if len(uri_set) == 1 and len(uri_set) == len(avail):
            verd, vfont = "Tutti concordano", Font(name="Arial",size=10,
                                                   color="1D6A38", bold=True)
        elif len(uri_set) == 1 and uri_set:
            verd, vfont = "Accordo parziale", Font(name="Arial",size=10,
                                                   color="85610A", bold=True)
        elif len(uri_set) > 1:
            verd, vfont = "Divergono", Font(name="Arial",size=10,
                                            color="A32D2D", bold=True)
        elif not uri_set and m1:
            verd, vfont = "Solo S1", Font(name="Arial",size=10,
                                          color="5C2D91", bold=True)
        else:
            verd, vfont = "Nessun match", Font(name="Arial",size=10,
                                               color="888888", italic=True)
        vc = ws1.cell(row=r_idx, column=verdict_col, value=verd)
        _style_cell(vc, fill=fill, font=vfont, wrap=False)

    ws1.row_dimensions[1].height = 30
    for r in range(2, len(db_jobs)+2):
        ws1.row_dimensions[r].height = 60

    # ── FOGLIO 2: Ontology Jobs ───────────────────────────────
    ws2 = wb.create_sheet("Ontology Jobs")
    ws2.freeze_panes = "A2"
    _set_header(ws2, 1, [
        (1,"#"),(2,"URI locale"),(3,"Label O*NET"),
        (4,"Net job titles (estratto)"),(5,"Descrizione O*NET"),
    ])
    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 48
    ws2.column_dimensions["C"].width = 36
    ws2.column_dimensions["D"].width = 50
    ws2.column_dimensions["E"].width = 65

    for i, j in enumerate(onto_jobs, 1):
        alt  = (i % 2 == 0)
        fill = _FILL_ALT if alt else None
        for col, val, wrap in [
            (1, i,               False),
            (2, j["local_name"], True),
            (3, j["label"],      True),
            (4, j["titles"][:300] if j["titles"] else "—", True),
            (5, j["description"][:500] if j["description"] else "—", True),
        ]:
            c = ws2.cell(row=i+1, column=col, value=val)
            _style_cell(c, fill=fill, wrap=wrap)
        ws2.row_dimensions[i+1].height = 60
    ws2.row_dimensions[1].height = 30

    # ── FOGLIO 3: Score Detail ────────────────────────────────
    ws3 = wb.create_sheet("Score Detail")
    ws3.freeze_panes = "C2"

    # Header: ID | Job Title | poi una colonna per ogni modello×professione
    hdr = [(1,"ID"),(2,"Job Title (DB)")]
    col = 3
    model_col_ranges: dict[str, tuple[int,int]] = {}
    for key in avail:
        start = col
        short = model_names.get(key, key).split("/")[-1]
        for j in onto_jobs:
            hdr.append((col, f"{short}\n{j['label'][:22]}"))
            col += 1
        model_col_ranges[key] = (start, col-1)
    _set_header(ws3, 1, hdr)
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 28
    for c in range(3, col):
        ws3.column_dimensions[get_column_letter(c)].width = 14

    # Righe dati con score raw per ogni combinazione
    for r_idx, row in enumerate(db_jobs, start=2):
        rid  = row["id"]
        ws3.cell(row=r_idx, column=1, value=rid)
        ws3.cell(row=r_idx, column=2, value=row["title"])
        for c in [1,2]:
            _style_cell(ws3.cell(row=r_idx, column=c),
                        fill=_FILL_ALT if r_idx%2==0 else None)

        for key in avail:
            r2       = nlp_idx[key].get(rid, {})
            m2       = r2.get("match")
            score_raw = r2.get("score_raw", 0.0)
            start_c, _ = model_col_ranges[key]

            # Per ora salviamo lo score best-match nella colonna della
            # professione trovata, e lasceremo vuote le altre.
            # (score completi richiederebbero di salvare l'intera matrice
            #  nel run_bert_matching — estensione futura)
            best_uri = m2["job"]["uri"] if m2 else None
            for c_off, j in enumerate(onto_jobs):
                c_abs = start_c + c_off
                is_best = (j["uri"] == best_uri)
                val = round(score_raw, 4) if is_best else None
                cell = ws3.cell(row=r_idx, column=c_abs, value=val)
                if is_best and val is not None:
                    fills = {"S2_MPNET":  _FILL_BERT,
                             "S4_MiniLM": _FILL_MINI}
                    _style_cell(cell, fill=fills.get(key),
                                font=_GREEN_FONT if score_raw >= 0.70
                                     else _BODY_FONT, wrap=False)
                else:
                    _style_cell(cell,
                                fill=_FILL_ALT if r_idx%2==0 else None,
                                wrap=False)

        ws3.row_dimensions[r_idx].height = 20
    ws3.row_dimensions[1].height = 45

    # ── FOGLIO 4: AI Judge ───────────────────────────────────
    if judge_results:
        _FILL_JUDGE_HDR  = PatternFill("solid", fgColor="3B1873")   # viola scuro
        _FONT_JUDGE_HDR  = Font(bold=True, color="FFFFFF", name="Arial", size=10)

        _FILL_CONFIRMED  = PatternFill("solid", fgColor="D4EDDA")   # verde chiaro
        _FILL_PARTIAL    = PatternFill("solid", fgColor="FFF3CD")   # giallo chiaro
        _FILL_REJECTED   = PatternFill("solid", fgColor="F8D7DA")   # rosso chiaro
        _FILL_UNCERTAIN  = PatternFill("solid", fgColor="F8F9FA")   # grigio chiaro

        _verdict_fill = {
            "CONFIRMED": _FILL_CONFIRMED,
            "PARTIAL":   _FILL_PARTIAL,
            "REJECTED":  _FILL_REJECTED,
            "UNCERTAIN": _FILL_UNCERTAIN,
        }
        _verdict_font_color = {
            "CONFIRMED": "1D6A38",
            "PARTIAL":   "85610A",
            "REJECTED":  "A32D2D",
            "UNCERTAIN": "888888",
        }

        ws4 = wb.create_sheet("AI Judge")
        ws4.freeze_panes = "A2"

        judge_hdr = [
            (1,  "ID"),
            (2,  "Job Title (DB)"),
            (3,  "S2 · Match O*NET"),
            (4,  "S2 · Score"),
            (5,  "S4 · Match O*NET"),
            (6,  "S4 · Score"),
            (7,  "Proposta principale"),
            (8,  "Verdetto Judge"),
            (9,  "Confidence"),
            (10, "Reasoning"),
            (11, "Suggested Label"),
            (12, "Latency (ms)"),
        ]
        for col_j, text_j in judge_hdr:
            c = ws4.cell(row=1, column=col_j, value=text_j)
            c.font      = _FONT_JUDGE_HDR
            c.fill      = PatternFill("solid", fgColor="3B1873")
            c.alignment = _CENTER
            c.border    = _BORDER

        ws4.column_dimensions["A"].width  = 8
        ws4.column_dimensions["B"].width  = 28
        ws4.column_dimensions["C"].width  = 30
        ws4.column_dimensions["D"].width  = 10
        ws4.column_dimensions["E"].width  = 30
        ws4.column_dimensions["F"].width  = 10
        ws4.column_dimensions["G"].width  = 30
        ws4.column_dimensions["H"].width  = 14
        ws4.column_dimensions["I"].width  = 12
        ws4.column_dimensions["J"].width  = 60
        ws4.column_dimensions["K"].width  = 30
        ws4.column_dimensions["L"].width  = 12

        for j_idx, jr in enumerate(judge_results, start=2):
            verdict  = jr.get("verdict", "UNCERTAIN")
            vfill    = _verdict_fill.get(verdict, _FILL_UNCERTAIN)
            vfc      = _verdict_font_color.get(verdict, "888888")
            vfont    = Font(name="Arial", size=10, color=vfc, bold=(verdict in ("CONFIRMED","REJECTED")))

            def wj(col_jj, val_jj, cfill=None, cfont=None, wrap=False):
                c = ws4.cell(row=j_idx, column=col_jj, value=val_jj)
                _style_cell(c, fill=cfill or (_FILL_ALT if j_idx % 2 == 0 else None),
                            font=cfont, wrap=wrap)

            wj(1,  jr.get("job_id",  ""),                        wrap=False)
            wj(2,  jr.get("job_title", ""),                       wrap=True)
            wj(3,  jr.get("s2_match") or "—",                    wrap=True)
            wj(4,  round(jr["s2_score"], 4) if jr.get("s2_score") else None, wrap=False)
            wj(5,  jr.get("s4_match") or "—",                    wrap=True)
            wj(6,  round(jr["s4_score"], 4) if jr.get("s4_score") else None, wrap=False)
            wj(7,  jr.get("primary_label", "—"),                 wrap=True)

            vc = ws4.cell(row=j_idx, column=8, value=verdict)
            _style_cell(vc, fill=vfill, font=vfont, wrap=False)

            cc = ws4.cell(row=j_idx, column=9,
                          value=round(jr.get("confidence", 0.0), 3))
            _style_cell(cc, fill=vfill, wrap=False)

            wj(10, jr.get("reasoning", ""),   wrap=True)
            wj(11, jr.get("suggested_label") or "—", wrap=True)
            wj(12, jr.get("latency_ms", 0),   wrap=False)

            ws4.row_dimensions[j_idx].height = 50

        ws4.row_dimensions[1].height = 30

    wb.save(path)
    return path


# ═════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Rientra@ Step 2 — Matching (no fine-tuning)")
    args = parser.parse_args()

    console.print()
    console.print(Panel(
        f"[bold white]Rientra@[/] — Step 2: RDB2RDF + String Matching + NLP\n"
        f"[dim]Modelli: MPNet · MiniLM  |  "
        f"Tutti sentence-transformers nativi  |  Metrica: cosine similarity[/]\n"
        f"[dim]{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}[/]",
        border_style="blue", padding=(0, 2),
    ))
    console.print()

    # ── [0] Setup ─────────────────────────────────────────────
    console.print(Rule("[bold cyan]0 · Setup[/]", style="cyan"))
    console.print()

    if not test_connection():
        sys.exit(1)

    with console.status(f"  Caricamento ontologia [cyan]{ONTOLOGY_FILE}[/]...",
                        spinner="dots"):
        onto_graph = load_ontology(ONTOLOGY_FILE)
    console.print(f"  [bold green]✓[/] Ontologia: [bold]{len(onto_graph)}[/] triple")

    onto_jobs = extract_ontology_jobs(onto_graph)
    console.print(f"  [bold green]✓[/] Professioni O*NET estratte: [bold]{len(onto_jobs)}[/]")
    console.print()
    print_ontology_table(onto_jobs)
    console.print()

    # ── [1] Lettura DB ────────────────────────────────────────
    console.print(Rule("[bold magenta]1 · Lettura DB esterno[/]", style="magenta"))
    console.print()
    db_jobs = fetch_all(
        "SELECT id, title, description FROM ext_job ORDER BY id")
    console.print(
        f"  [bold green]✓[/] [bold]{len(db_jobs)}[/] lavori da [cyan]ext_job[/]")
    console.print()
    print_db_table(db_jobs)
    console.print()

    # ── [2] Strategia 1 — String Matching ─────────────────────
    console.print(Rule("[bold yellow]2 · Strategia 1 — String Matching (pre-processing esteso)[/]",
                       style="yellow"))
    console.print(
        "  [dim]Jaccard (0.5) + Overlap (0.3) + Levenshtein (0.2)[/]\n"
        "  [dim]Pre-processing: lowercase · simboli → spazio · "
        "espansione abbreviazioni · rimozione stopwords[/]")
    console.print()
    s1_res = run_string_matching(db_jobs, onto_jobs)
    console.print(f"  [bold green]✓[/] Completato\n")

    # ── [3] Strategia 2 — MPNet sentence-transformer ──────────
    console.print(Rule("[bold blue]3 · Strategia 2 — all-mpnet-base-v2[/]",
                       style="blue"))
    console.print(
        f"  [dim]Modello: {MODELS['S2_MPNET']}[/]\n"
        f"  [dim]Sentence-transformer nativo · 768 dim · Soglia: {MPNET_THRESHOLD}[/]\n"
        f"  [dim]Ottimizzato per testi lunghi · nessun problema di anisotropy[/]")
    console.print()
    s2_res, s2_name = run_st_matching(
        db_jobs, onto_jobs,
        model_name_or_path=MODELS["S2_MPNET"],
        strategy_key="S2_MPNET",
        threshold=MPNET_THRESHOLD,
    )
    if s2_res:
        console.print(f"  [bold green]✓[/] Completato\n")
    else:
        console.print(
            f"  [yellow]⚠[/] Saltato — installa [cyan]sentence-transformers[/]\n")

    # ── [4] Strategia 4 — MiniLM ──────────────────────────────
    console.print(Rule("[bold green]4 · Strategia 4 — MiniLM (riferimento)[/]",
                       style="green"))
    console.print(
        f"  [dim]Modello: {MODELS['S4_MiniLM']}  |  "
        f"Fine-tuned per sentence similarity  |  Soglia: {MINILM_THRESHOLD}[/]")
    console.print()
    s4_res, s4_name = run_minilm_matching(
        db_jobs, onto_jobs, threshold=MINILM_THRESHOLD)
    if s4_res:
        console.print(f"  [bold green]✓[/] Completato\n")
    else:
        console.print(
            f"  [yellow]⚠[/] Saltato — installa [cyan]sentence-transformers[/]\n")

    # Raccoglie risultati disponibili
    nlp_results: dict[str, list[dict]] = {}
    model_names: dict[str, str]        = {}
    if s2_res:
        nlp_results["S2_MPNET"] = s2_res
        model_names["S2_MPNET"] = s2_name
    if s4_res:
        nlp_results["S4_MiniLM"] = s4_res
        model_names["S4_MiniLM"] = s4_name

    # ── [5] Confronto ─────────────────────────────────────────
    console.print(Rule("[bold white]5 · Confronto strategie[/]", style="white"))
    console.print()
    if nlp_results:
        print_full_comparison(s1_res, nlp_results, model_names)
        print_summary_multi(s1_res, nlp_results, model_names)
    else:
        console.print(
            "  [yellow]Nessun modello NLP disponibile — solo S1[/]")
    console.print()

    # ── [6] AI Judge — validazione locale Ollama ───────────────
    from ai_judge import (
        check_ollama_available, run_ai_judge, print_judge_table
    )

    console.print(Rule(
        f"[bold magenta]6 · AI Judge — {OLLAMA_MODEL}[/]",
        style="magenta"
    ))
    console.print(
        f"  [dim]Modello locale: [cyan]{OLLAMA_MODEL}[/] via Ollama[/]\n"
        f"  [dim]Valida semanticamente i match di S2 e S4 → CONFIRMED / PARTIAL / REJECTED / UNCERTAIN[/]"
    )
    console.print()

    judge_results: list[dict] | None = None
    if check_ollama_available(OLLAMA_MODEL) and nlp_results:
        console.print(
            f"  [bold green]✓[/] Ollama pronto — avvio judge su "
            f"[bold]{len(db_jobs)}[/] job...\n"
        )
        judge_results = run_ai_judge(
            db_jobs, nlp_results, onto_jobs, model=OLLAMA_MODEL
        )
        console.print(f"  [bold green]✓[/] Judge completato\n")
        print_judge_table(judge_results, model=OLLAMA_MODEL)
    elif not nlp_results:
        console.print(
            "  [yellow]⚠[/] Nessun risultato NLP disponibile — judge saltato\n"
        )
    else:
        console.print(
            f"  [yellow]⚠[/] Ollama non disponibile — step 6 saltato\n"
            f"  [dim]Installa Ollama da https://ollama.com/download, poi:[/]\n"
            f"  [dim]  ollama pull {OLLAMA_MODEL}[/]\n"
        )
    console.print()

    # ── [6] Output Excel ──────────────────────────────────────
    console.print(Rule("[bold blue]6 · Produzione file Excel[/]", style="blue"))
    console.print()
    with console.status("  Costruzione workbook...", spinner="dots"):
        xlsx_path = save_excel(db_jobs, s1_res, nlp_results,
                               model_names, onto_jobs,
                               judge_results=judge_results)

    size_kb = os.path.getsize(xlsx_path) // 1024
    t = Table(box=box.SIMPLE, show_header=False, padding=(0, 1))
    t.add_column(style="dim")
    t.add_column(style="cyan")
    t.add_column(style="dim", justify="right")
    t.add_row("Excel",    xlsx_path,           f"{size_kb:,} KB")
    t.add_row("Foglio 1", "Matching Results",
              f"{len(db_jobs)} righe · tutte le strategie")
    t.add_row("Foglio 2", "Ontology Jobs",
              f"{len(onto_jobs)} professioni O*NET")
    t.add_row("Foglio 3", "Score Detail",
              "score per job × modello")
    if judge_results:
        t.add_row("Foglio 4", "AI Judge",
                  f"{len(judge_results)} verdetti · {OLLAMA_MODEL}")
    console.print(t)
    console.print()

    # ── Fine ──────────────────────────────────────────────────
    console.print(Panel(
        f"[bold green]Step 2 completato.[/]\n\n"
        f"Output → [cyan]{xlsx_path}[/]\n\n"
        f"[dim]Modelli NLP: "
        f"{', '.join(model_names.values()) or 'solo S1'}[/]\n"
        f"[dim]AI Judge: {OLLAMA_MODEL if judge_results else 'non eseguito'}[/]\n"
        f"[dim]Prossimo step: ESCO come ponte ontologico[/]",
        border_style="green", padding=(0, 2),
    ))
    console.print()


if __name__ == "__main__":
    main()