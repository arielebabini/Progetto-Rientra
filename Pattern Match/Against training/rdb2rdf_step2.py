"""
rdb2rdf_step2.py
================
Step 2 del progetto Rientra@: RDB2RDF con matching verso l'ontologia principale.

Dataset in ingresso:
  - Tabella PostgreSQL  : ext_job  (ext01…ext09)
  - Ontologia           : Rientra.rdf  (namespace http://www.stiima.cnr.it/JobList#)
  - Dati di training    : bert_training_data.xlsx  (60 lavori, 720 coppie)

Strategie di matching:
  S1 — String Matching          : Jaccard + Overlap + Levenshtein
                                   Pre-processing esteso: normalizzazione simboli,
                                   espansione abbreviazioni, rimozione stopwords.
  S2 — all-mpnet-base-v2        : sentence-transformer nativo (768 dim), cosine similarity.
                                   Superiore a MiniLM su testi lunghi. Nessun problema
                                   di anisotropy perché ottimizzato per sentence similarity.
  S3 — MPNet fine-tuned         : all-mpnet-base-v2 fine-tuned su bert_training_data.xlsx
                                   con CosineSimilarityLoss (architettura Siamese).
                                   Specializzato sul dominio Rientra@.
  S4 — all-MiniLM-L6-v2        : sentence-transformer compatto (384 dim), riferimento.

Nota sui modelli:
  Tutti i modelli NLP usati (S2, S3, S4) sono sentence-transformers nativi,
  non encoder BERT raw. Questo garantisce embedding di frase calibrati per
  cosine similarity e assenza del fenomeno di anisotropy.
  BERT base raw è stato rimosso in seguito alle osservazioni del referente.

Requisiti:
    pip install psycopg2-binary rdflib rich sentence-transformers openpyxl

Uso:
    python rdb2rdf_step2.py
    python rdb2rdf_step2.py --skip-training   # riusa modello già salvato
"""

import os
import sys
import numpy as np
from datetime import datetime

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich import box
from rich.text import Text
from rich.rule import Rule

console = Console()

# ─────────────────────────────────────────────────────────────
# CONFIGURAZIONE
# ─────────────────────────────────────────────────────────────

DB_CONFIG = {
    "host":     "localhost",
    "port":     5432,
    "dbname":   "rientra_db",
    "user":     "postgres",
    "password": "postgres",
}

ONTOLOGY_FILE    = "Rientra.rdf"
OUTPUT_DIR       = "output"
TRAINING_FILE    = "bert_training_data.xlsx"
FINETUNED_DIR    = os.path.join("output", "mpnet_finetuned")  # modello fine-tuned salvato

JOBLIST_BASE = "http://www.stiima.cnr.it/JobList#"
RIENTRA_BASE = "https://www.stiima.cnr.it/rientra#"

# Soglie cosine similarity
# Tutti i modelli NLP sono sentence-transformers → stessa scala di score
S1_THRESHOLD        = 0.12
MPNET_THRESHOLD     = 0.50   # all-mpnet-base-v2: sentence-transformer nativo
FINETUNED_THRESHOLD = 0.50   # MPNet fine-tuned: stessa scala
MINILM_THRESHOLD    = 0.50   # all-MiniLM-L6-v2: riferimento

# Iperparametri fine-tuning
FT_EPOCHS       = 4
FT_BATCH_SIZE   = 16
FT_LR           = 2e-5
FT_WARMUP_STEPS = 50

# Modelli — tutti sentence-transformers nativi
MODELS = {
    "S2_MPNET":    "sentence-transformers/all-mpnet-base-v2",
    "S3_FINETUNED": FINETUNED_DIR,   # MPNet fine-tuned su dominio Rientra@
    "S4_MiniLM":   "all-MiniLM-L6-v2",
}

# Abbreviazioni comuni nel dominio IT/lavoro da espandere nel pre-processing S1
ABBREVIATIONS = {
    "ml":   "machine learning",
    "ai":   "artificial intelligence",
    "nlp":  "natural language processing",
    "hr":   "human resources",
    "pr":   "public relations",
    "it":   "information technology",
    "db":   "database",
    "mgr":  "manager",
    "mgmt": "management",
    "admin": "administration",
    "asst": "assistant",
    "assoc": "associate",
    "spec": "specialist",
    "coord": "coordinator",
    "dept": "department",
    "exec": "executive",
    "sr":   "senior",
    "jr":   "junior",
    "r&d":  "research and development",
    "b2b":  "business to business",
    "b2c":  "business to consumer",
    "erp":  "enterprise resource planning",
    "crm":  "customer relationship management",
    "ui":   "user interface",
    "ux":   "user experience",
    "qa":   "quality assurance",
    "ops":  "operations",
    "dev":  "developer",
    "sw":   "software",
    "hw":   "hardware",
}


# ─────────────────────────────────────────────────────────────
# HELPERS VISIVI
# ─────────────────────────────────────────────────────────────

def score_color(score: float) -> str:
    if score >= 0.85: return "bold green"
    if score >= 0.70: return "green"
    if score >= 0.50: return "yellow"
    if score >= 0.30: return "orange3"
    return "red"

def score_bar(score: float, width: int = 8) -> str:
    filled = round(score * width)
    return "█" * filled + "░" * (width - filled)


# ═════════════════════════════════════════════════════════════
# UTILITÀ DB
# ═════════════════════════════════════════════════════════════

def get_connection():
    return psycopg2.connect(**DB_CONFIG)

def fetch_all(sql: str, params=None) -> list[dict]:
    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            return [dict(r) for r in cur.fetchall()]

def test_connection() -> bool:
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT version();")
                v = cur.fetchone()[0]
        console.print(f"  [bold green]✓[/] PostgreSQL [dim]{v[:60]}...[/]")
        return True
    except Exception as e:
        console.print(f"  [bold red]✗[/] Errore connessione: [red]{e}[/]")
        return False


# ═════════════════════════════════════════════════════════════
# ESTRAZIONE VOCABOLARIO DALL'ONTOLOGIA
# ═════════════════════════════════════════════════════════════

def load_ontology(path: str):
    """Carica il file RDF e restituisce un oggetto Graph rdflib."""
    from rdflib import Graph
    g = Graph()
    g.parse(path, format="xml")
    return g

def extract_ontology_jobs(onto_graph) -> list[dict]:
    from rdflib import RDF, OWL
    from rdflib.namespace import RDFS
    jobs = []
    for s, _, _ in onto_graph.triples((None, RDF.type, OWL.Class)):
        if not str(s).startswith(JOBLIST_BASE):
            continue
        local = str(s).split("#")[1]
        if local in ("Job", "Job_Descriptor"):
            continue

        label = str(onto_graph.value(s, RDFS.label) or "").strip()
        titles, desc = "", ""
        for pred, obj in onto_graph.predicate_objects(s):
            ps = str(pred)
            if "Net_job_titles" in ps or "job_titles" in ps:
                titles = str(obj)
            if "Net_short_description" in ps or "short_description" in ps:
                desc = str(obj)

        if titles and ":" in titles:
            titles = titles.split(":", 1)[1].strip()

        display = label or local.replace("_", " ")
        jobs.append({
            "uri":         str(s),
            "local_name":  local,
            "label":       display,
            "titles":      titles,
            "description": desc,
        })

    jobs.sort(key=lambda x: x["label"])
    return jobs

def _build_onto_text(j: dict) -> str:
    """Testo composito per l'ontologia: label + titoli alternativi + descrizione."""
    parts = [j["label"]]
    if j["titles"]:
        parts.append(j["titles"])
    if j["description"]:
        parts.append(j["description"])
    return " ".join(parts)


# ═════════════════════════════════════════════════════════════
# STRATEGIA 1 — STRING MATCHING (pre-processing esteso)
# ═════════════════════════════════════════════════════════════

# Stopwords leggere per il dominio job matching
# (parole molto frequenti che non aggiungono discriminazione)
_STOPWORDS = {
    "and", "or", "the", "a", "an", "of", "in", "for", "to",
    "with", "at", "by", "as", "on", "is", "are", "be", "been",
    "from", "that", "this", "which", "who", "other", "related",
}

def _norm(s: str) -> str:
    """
    Pre-processing esteso per string matching.

    Passi applicati in ordine:
      1. Lowercase e strip degli spazi
      2. Sostituzione dei separatori (-, /, _, .) con spazio
      3. Rimozione della punteggiatura residua (',', '(', ')', ecc.)
      4. Espansione delle abbreviazioni comuni (ML → machine learning)
      5. Rimozione delle stopwords leggere
      6. Collasso degli spazi multipli
    """
    import re
    s = s.lower().strip()
    # Separatori → spazio
    s = re.sub(r"[-/_.]", " ", s)
    # Punteggiatura residua
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    # Espansione abbreviazioni (parola intera, non sottostringa)
    tokens = s.split()
    expanded = []
    for tok in tokens:
        expanded.append(ABBREVIATIONS.get(tok, tok))
    # Rimozione stopwords + collasso spazi
    tokens = " ".join(expanded).split()
    tokens = [t for t in tokens if t not in _STOPWORDS and len(t) > 1]
    return " ".join(tokens)

def _jaccard(a: str, b: str) -> float:
    a, b = set(_norm(a).split()), set(_norm(b).split())
    return len(a & b) / len(a | b) if (a | b) else 1.0

def _overlap(a: str, b: str) -> float:
    a, b = set(_norm(a).split()), set(_norm(b).split())
    return len(a & b) / min(len(a), len(b)) if (a and b) else 0.0

def _levenshtein(s1: str, s2: str) -> float:
    s1, s2 = _norm(s1), _norm(s2)
    m, n = len(s1), len(s2)
    if not m and not n:
        return 1.0
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev, dp[0] = dp[:], i
        for j in range(1, n + 1):
            cost = 0 if s1[i-1] == s2[j-1] else 1
            dp[j] = min(prev[j] + 1, dp[j-1] + 1, prev[j-1] + cost)
    return 1.0 - dp[n] / max(m, n)

def _combined(a: str, b: str) -> float:
    return 0.5 * _jaccard(a, b) + 0.3 * _overlap(a, b) + 0.2 * _levenshtein(a, b)

def s1_match_one(title: str, onto_jobs: list[dict]) -> dict | None:
    best, bj, bv = 0.0, None, None
    for j in onto_jobs:
        cands = [j["label"]]
        if j["titles"]:
            cands += [t.strip() for t in j["titles"].split(",") if t.strip()]
        for c in cands:
            sc = _combined(title, c)
            if sc > best:
                best, bj, bv = sc, j, c
    if best >= S1_THRESHOLD:
        return {"job": bj, "score": best, "via": bv}
    return None

def run_string_matching(db_jobs: list[dict], onto_jobs: list[dict]) -> list[dict]:
    results = []
    with Progress(SpinnerColumn(), TextColumn("{task.description}"),
                  BarColumn(), TaskProgressColumn(),
                  console=console, transient=True) as p:
        task = p.add_task("  Calcolo...", total=len(db_jobs))
        for r in db_jobs:
            results.append({"id": r["id"], "title": r["title"],
                            "match_s1": s1_match_one(r["title"], onto_jobs)})
            p.advance(task)
    return results


# ═════════════════════════════════════════════════════════════
# COSINE SIMILARITY — funzione centrale condivisa da S2 e S3
# ═════════════════════════════════════════════════════════════

def cosine_similarity_matrix(query_emb: np.ndarray,
                              corpus_embs: np.ndarray) -> np.ndarray:
    """
    Calcola la cosine similarity tra un vettore query e una matrice di vettori corpus.

    Formula:
        cos(θ) = (A · B) / (‖A‖ · ‖B‖)

    dove A è il vettore del job nel DB e B è il vettore di una professione O*NET.

    Il risultato è compreso tra -1 e 1:
      1.0  → testi identici nel significato
      0.0  → testi ortogonali (nessuna relazione semantica)
     -1.0  → testi opposti (raro con embedding linguistici)

    Per embedding BERT non negativi il range effettivo è [0, 1].
    """
    query_norm  = query_emb / (np.linalg.norm(query_emb) + 1e-9)
    corpus_norm = corpus_embs / (np.linalg.norm(corpus_embs, axis=1, keepdims=True) + 1e-9)
    return corpus_norm.dot(query_norm)


# ═════════════════════════════════════════════════════════════
# EMBEDDING CON BERT RAW (BERT base e BioBERT)
# ═════════════════════════════════════════════════════════════

# ═════════════════════════════════════════════════════════════
# SENTENCE-TRANSFORMER MATCHING — funzione generica per S2, S3, S4
# ═════════════════════════════════════════════════════════════

def run_st_matching(db_jobs: list[dict],
                    onto_jobs: list[dict],
                    model_name_or_path: str,
                    strategy_key: str,
                    threshold: float) -> tuple[list[dict], str]:
    """
    Esegue il matching con un qualsiasi modello sentence-transformers.

    Usato per S2 (all-mpnet-base-v2), S3 (MPNet fine-tuned) e S4 (MiniLM).
    Tutti questi modelli sono sentence-transformers nativi: il loro .encode()
    produce embedding ottimizzati per cosine similarity senza bisogno di
    estrarre manualmente il token [CLS] o applicare pooling.

    La cosine similarity è calcolata esplicitamente con cosine_similarity_matrix
    per uniformità metodologica tra tutte le strategie NLP.
    """
    try:
        from sentence_transformers import SentenceTransformer

        label = model_name_or_path.split("/")[-1]
        with console.status(
                f"  Caricamento [cyan]{label}[/]...", spinner="dots"):
            model = SentenceTransformer(model_name_or_path)
        console.print(
            f"  [bold green]✓[/] Modello caricato: [cyan]{model_name_or_path}[/]")

        onto_texts = [_build_onto_text(j) for j in onto_jobs]
        db_texts   = [f"{r['title']}. {r.get('description', '')}"
                      for r in db_jobs]

        with Progress(SpinnerColumn(), TextColumn("{task.description}"),
                      BarColumn(), TaskProgressColumn(),
                      console=console, transient=True) as p:
            t1 = p.add_task("  Encoding O*NET...", total=1)
            onto_embs = model.encode(onto_texts, convert_to_numpy=True)
            p.advance(t1)

            results = []
            t2 = p.add_task("  Cosine similarity...", total=len(db_jobs))
            for i, row in enumerate(db_jobs):
                q_emb = model.encode(db_texts[i], convert_to_numpy=True)
                sims  = cosine_similarity_matrix(q_emb, onto_embs)
                idx   = int(np.argmax(sims))
                sc    = float(sims[idx])
                m = {"job": onto_jobs[idx], "score": sc} \
                    if sc >= threshold else None
                results.append({
                    "id":        row["id"],
                    "title":     row["title"],
                    "match":     m,
                    "score_raw": sc,
                    "strategy":  strategy_key,
                })
                p.advance(t2)

        return results, model_name_or_path

    except Exception as e:
        console.print(
            f"  [yellow]⚠[/] {model_name_or_path} non disponibile: [dim]{e}[/]")
        return None, None


# ═════════════════════════════════════════════════════════════
# FINE-TUNING BERT (S3)
# ═════════════════════════════════════════════════════════════

def load_training_pairs(xlsx_path: str) -> list[tuple[str, str, float]]:
    """
    Legge il file Excel di training e restituisce le coppie per il fine-tuning.

    Logica di lettura della colonna Validated (L):
    - Se l'utente ha scritto SI a mano per ALCUNE righe e NO per ALTRE
      in modo non proporzionale al label → modalita' revisione manuale
      (usa solo le righe con SI, esclude quelle con NO)
    - In tutti gli altri casi (colonna vuota, formule automatiche, o
      SI/NO che ricalcano esattamente il label) → usa tutto il dataset,
      sia positive che negative, che e' il comportamento corretto per
      il fine-tuning con CosineSimilarityLoss.

    Il training richiede SEMPRE sia positive che negative.
    """
    from openpyxl import load_workbook

    wb   = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws   = wb["Training Pairs"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    all_rows = []
    for row in rows:
        if len(row) < 6 or row[3] is None:
            continue
        text_a    = str(row[3]).strip()
        text_b    = str(row[4]).strip() if row[4] else ""
        label_raw = row[5]
        meaning   = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        valid_raw = str(row[11]).strip() if len(row) > 11 and row[11] else ""

        if not text_a or not text_b:
            continue

        # Label numerico dalla colonna F o G
        if label_raw is not None:
            try:
                label = float(label_raw)
            except (ValueError, TypeError):
                label = 1.0 if "stesso" in str(label_raw).lower() else 0.0
        elif meaning:
            label = 1.0 if "stesso" in meaning.lower() else 0.0
        else:
            continue

        valid_upper = valid_raw.upper().strip("'")
        all_rows.append((text_a, text_b, label, valid_upper))

    if not all_rows:
        console.print("  [bold red]✗[/] Nessuna coppia trovata nel file.")
        return []

    # Analizza il pattern della colonna Validated
    si_rows  = [(a,b,l) for a,b,l,v in all_rows if v == "SI"]
    no_rows  = [(a,b,l) for a,b,l,v in all_rows if v == "NO"]
    all_pairs = [(a,b,l) for a,b,l,v in all_rows]

    si_pos = sum(1 for _,_,l in si_rows if l == 1.0)
    si_neg = sum(1 for _,_,l in si_rows if l == 0.0)
    no_pos = sum(1 for _,_,l in no_rows if l == 1.0)
    no_neg = sum(1 for _,_,l in no_rows if l == 0.0)

    # Rileva se la colonna Validated ricalca esattamente il label
    # (formula automatica: SI=positivo, NO=negativo)
    formula_pattern = (si_neg == 0 and no_pos == 0 and
                       len(si_rows) > 0 and len(no_rows) > 0)

    if formula_pattern:
        # La colonna Validated e' una formula automatica che non aggiunge
        # informazione rispetto al label → usa tutto il dataset
        pairs = all_pairs
        console.print(
            f"  [dim]Colonna Validated rispecchia il label (formula automatica) "
            f"— usate tutte le [bold]{len(pairs)}[/] coppie[/]"
        )
    elif si_rows and (si_neg > 0 or no_pos > 0):
        # L'utente ha validato manualmente in modo non correlato al label
        pairs = si_rows
        console.print(
            f"  [bold green]✓[/] Revisione manuale: "
            f"[bold]{len(pairs)}[/] coppie selezionate con Validated=SI"
        )
    elif si_rows and si_neg == 0 and no_rows and no_pos > 0:
        # Altro caso di validazione parziale
        pairs = si_rows
        console.print(
            f"  [bold green]✓[/] Usate [bold]{len(pairs)}[/] coppie (Validated=SI)"
        )
    else:
        # Nessuna validazione → usa tutto
        pairs = all_pairs
        console.print(
            f"  [dim]Nessuna validazione esplicita — "
            f"usate tutte le [bold]{len(pairs)}[/] coppie[/]"
        )

    n_pos = sum(1 for _, _, l in pairs if l == 1.0)
    n_neg = len(pairs) - n_pos

    if n_pos == 0 or n_neg == 0:
        console.print(
            f"  [bold yellow]⚠[/] Solo coppie "
            f"{'positive' if n_pos > 0 else 'negative'} — "
            f"aggiunto il dataset completo per bilanciare."
        )
        # Fallback: usa tutto il dataset per garantire il bilanciamento
        pairs = all_pairs
        n_pos = sum(1 for _, _, l in pairs if l == 1.0)
        n_neg = len(pairs) - n_pos

    console.print(
        f"  [dim]Positive (label=1): {n_pos}  |  Negative (label=0): {n_neg}[/]"
    )
    return pairs



def finetune_mpnet(training_pairs: list[tuple[str, str, float]],
                   save_dir: str) -> str:
    """
    Esegue il fine-tuning di all-mpnet-base-v2 su coppie (text_a, text_b, label)
    con CosineSimilarityLoss tramite sentence-transformers.

    Perché MPNet invece di BERT base:
      - MPNet è già un sentence-transformer nativo: produce embedding calibrati
        per cosine similarity senza anisotropy.
      - Il fine-tuning parte da uno spazio vettoriale già ben strutturato,
        quindi converge più velocemente e produce risultati più affidabili.
      - Gli embedding hanno 768 dimensioni (vs 384 di MiniLM): più capacità
        rappresentativa, utile per testi lunghi come le descrizioni di lavoro.

    Architettura Siamese con CosineSimilarityLoss:
      - Due passaggi del modello (pesi condivisi) producono emb_A e emb_B.
      - loss = MSE(cosine_sim(emb_A, emb_B), label)
      - label=1 → spinge cosine_sim verso 1 (vettori paralleli)
      - label=0 → spinge cosine_sim verso 0 (vettori ortogonali)

    Il modello viene salvato in save_dir e ricaricato nelle esecuzioni
    successive con --skip-training.
    """
    try:
        from sentence_transformers import (SentenceTransformer,
                                           InputExample, losses)
        from torch.utils.data import DataLoader
        import math

        base_model = MODELS["S2_MPNET"]
        console.print(f"  [dim]Base model: {base_model}[/]")
        console.print(
            f"  [dim]Epoche: {FT_EPOCHS}  Batch: {FT_BATCH_SIZE}"
            f"  LR: {FT_LR}  Warmup: {FT_WARMUP_STEPS}[/]")

        with console.status(
                f"  Caricamento {base_model.split('/')[-1]}...",
                spinner="dots"):
            model = SentenceTransformer(base_model)

        examples = [
            InputExample(texts=[a, b], label=float(l))
            for a, b, l in training_pairs
        ]
        loader  = DataLoader(examples, shuffle=True,
                             batch_size=FT_BATCH_SIZE)
        loss_fn = losses.CosineSimilarityLoss(model)

        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console,
        ) as p:
            task = p.add_task(
                f"  Fine-tuning MPNet "
                f"({FT_EPOCHS} epoche, {len(examples)} coppie)...",
                total=FT_EPOCHS,
            )
            model.fit(
                train_objectives=[(loader, loss_fn)],
                epochs=FT_EPOCHS,
                warmup_steps=FT_WARMUP_STEPS,
                optimizer_params={"lr": FT_LR},
                show_progress_bar=False,
                callback=lambda score, epoch, steps: p.advance(task),
            )

        os.makedirs(save_dir, exist_ok=True)
        model.save(save_dir)
        console.print(
            f"  [bold green]✓[/] Modello salvato in: [cyan]{save_dir}[/]")
        return save_dir

    except Exception as e:
        console.print(f"  [bold red]✗[/] Fine-tuning fallito: [red]{e}[/]")
        import traceback; traceback.print_exc()
        return None


def run_finetuned_matching(db_jobs: list[dict],
                           onto_jobs: list[dict],
                           model_path: str,
                           threshold: float) -> tuple[list[dict], str]:
    """
    Matching con il modello MPNet fine-tuned su Rientra@.
    Delega a run_st_matching con strategy_key S3_FINETUNED.
    """
    res, name = run_st_matching(
        db_jobs, onto_jobs, model_path, "S3_FINETUNED", threshold)
    if res:
        # Rinomina il modello per chiarezza nell'output
        return res, "all-mpnet-base-v2 (fine-tuned Rientra@)"
    return None, None


# ═════════════════════════════════════════════════════════════
# EMBEDDING CON SENTENCE-TRANSFORMERS (MiniLM — S4)
# ═════════════════════════════════════════════════════════════

def run_minilm_matching(db_jobs: list[dict],
                        onto_jobs: list[dict],
                        threshold: float) -> tuple[list[dict], str]:
    """
    Matching con all-MiniLM-L6-v2 (modello di riferimento compatto).
    Delega a run_st_matching con strategy_key S4_MiniLM.
    """
    return run_st_matching(
        db_jobs, onto_jobs,
        MODELS["S4_MiniLM"], "S4_MiniLM", threshold)


# ═════════════════════════════════════════════════════════════
# OUTPUT RICH
# ═════════════════════════════════════════════════════════════

def print_ontology_table(onto_jobs: list[dict]):
    t = Table(title="Professioni O*NET nell'ontologia",
              box=box.ROUNDED, header_style="bold cyan",
              title_style="bold white", show_lines=False)
    t.add_column("#",           style="dim",       width=3,  justify="right")
    t.add_column("URI locale",  style="cyan",       width=46)
    t.add_column("Label O*NET", style="bold white", width=38)
    t.add_column("Desc.",       justify="center",   width=6)
    for i, j in enumerate(onto_jobs, 1):
        t.add_row(str(i), j["local_name"], j["label"],
                  "[green]✓[/]" if j["description"] else "[red]✗[/]")
    console.print(t)


def print_db_table(db_jobs: list[dict]):
    t = Table(title="Lavori nel DB esterno (ext_job)",
              box=box.ROUNDED, header_style="bold magenta",
              title_style="bold white", show_lines=False)
    t.add_column("ID",    style="bold magenta", width=6)
    t.add_column("Titolo",                      width=30)
    t.add_column("Descrizione (estratto)",      width=62)
    for r in db_jobs:
        desc = (r.get("description") or "")[:60] + "…"
        t.add_row(r["id"], r["title"], f"[dim]{desc}[/]")
    console.print(t)


def _score_cell(score: float | None, raw: float = 0.0) -> tuple[Text, Text]:
    """Restituisce (cella_label_aggiuntiva, cella_score) per la tabella."""
    if score is not None:
        bar = Text()
        bar.append(f"{score:.3f}\n", style=score_color(score))
        bar.append(score_bar(score),  style=score_color(score))
        return Text(""), bar
    else:
        lbl = Text(f"(raw: {raw:.3f})", style="dim red")
        bar = Text()
        bar.append(f"{raw:.3f}\n", style="dim red")
        bar.append(score_bar(raw),  style="dim red")
        return lbl, bar


def print_full_comparison(s1_res: list[dict],
                          nlp_results: dict[str, list[dict]],
                          model_names: dict[str, str]):
    """
    Tabella unica con S1 + tutti i modelli NLP disponibili, una colonna per modello.
    nlp_results: { "S2_MPNET": [...], "S3_FINETUNED": [...], "S4_MiniLM": [...] }
    model_names: { "S2_MPNET": "all-mpnet-base-v2", ... }
    """
    # Indici per lookup rapido per ogni strategia
    idx = {key: {r["id"]: r for r in res}
           for key, res in nlp_results.items() if res}

    available_strategies = [k for k in ["S2_MPNET", "S3_FINETUNED", "S4_MiniLM"]
                            if k in idx]

    # Costruisce header dinamico
    title_parts = ["S1 String"] + [
        model_names.get(k, k).split("/")[-1][:18]
        for k in available_strategies
    ]
    t = Table(
        title="Confronto strategie di matching · cosine similarity",
        box=box.SIMPLE_HEAD,
        show_lines=True,
        header_style="bold white on dark_blue",
        title_style="bold white",
        expand=True,
    )
    t.add_column("ID",       style="bold", width=6,  justify="center")
    t.add_column("Titolo DB",              width=26)
    t.add_column("S1 — Match",             width=28)
    t.add_column("Score S1", justify="center", width=14)

    for k in available_strategies:
        short = model_names.get(k, k).split("/")[-1][:20]
        t.add_column(f"{k[:2]} — {short}", width=28)
        t.add_column(f"Score {k[:2]}", justify="center", width=14)

    t.add_column("Verdetto", justify="center", width=10)

    for r in s1_res:
        rid = r["id"]
        m1  = r.get("match_s1")

        # S1
        if m1:
            s1_cell = Text()
            s1_cell.append(m1["job"]["label"] + "\n")
            s1_cell.append(f"via: {m1['via'][:24]}", style="dim")
            s1_sc = Text()
            s1_sc.append(f"{m1['score']:.3f}\n", style=score_color(m1["score"]))
            s1_sc.append(score_bar(m1["score"]),  style=score_color(m1["score"]))
        else:
            s1_cell = Text("— nessun match —", style="dim")
            s1_sc   = Text("—", style="dim")

        row_cells = [rid, r["title"], s1_cell, s1_sc]

        # NLP strategies
        nlp_matches = []
        for k in available_strategies:
            r2   = idx[k].get(rid, {})
            m2   = r2.get("match")
            raw  = r2.get("score_raw", 0.0)
            if m2:
                cell = Text(m2["job"]["label"])
                sc   = Text()
                sc.append(f"{m2['score']:.3f}\n", style=score_color(m2["score"]))
                sc.append(score_bar(m2["score"]),  style=score_color(m2["score"]))
                nlp_matches.append(m2["job"]["uri"])
            else:
                cell = Text()
                cell.append("— sotto soglia —\n", style="dim")
                cell.append(f"(raw: {raw:.3f})", style="dim red")
                sc   = Text()
                sc.append(f"{raw:.3f}\n", style="dim red")
                sc.append(score_bar(raw),  style="dim red")
                nlp_matches.append(None)
            row_cells += [cell, sc]

        # Verdetto: consensus tra i modelli che hanno trovato un match
        valid = [u for u in nlp_matches if u]
        if not valid and not m1:
            verdetto = Text("✗ no", style="bold red")
        elif len(set(valid)) == 1 and len(valid) == len(available_strategies):
            verdetto = Text("✓ tutti", style="bold green")
        elif len(set(valid)) == 1 and valid:
            verdetto = Text("~ parz.", style="yellow")
        elif valid:
            verdetto = Text("≠ div.", style="orange3")
        else:
            verdetto = Text("S1 only", style="magenta")

        row_cells.append(verdetto)
        t.add_row(*row_cells)

    console.print(t)
    console.print(
        "  [bold green]✓ tutti[/] tutti i modelli NLP concordano  "
        "[yellow]~ parz.[/] accordo parziale  "
        "[orange3]≠ div.[/] modelli divergono  "
        "[magenta]S1 only[/] solo string match  "
        "[bold red]✗ no[/] nessun match\n"
    )


def print_cosine_detail(nlp_results: dict[str, list[dict]],
                        model_names: dict[str, str],
                        onto_jobs: list[dict]):
    """
    Tabella dettaglio cosine similarity: per ogni job DB mostra lo score
    con TUTTE le professioni O*NET, per ogni modello disponibile.
    Utile per capire la distribuzione degli score e validare le soglie.
    """
    console.print(Rule("[dim]Dettaglio score cosine per tutti i candidati O*NET[/]", style="dim"))
    console.print("[dim]  (i valori in verde sono quelli scelti come match migliore)[/]\n")

    available = [k for k in ["S2_MPNET", "S3_FINETUNED", "S4_MiniLM"]
                 if k in nlp_results and nlp_results[k]]

    for strategy_key in available:
        res_list = nlp_results[strategy_key]
        mname    = model_names.get(strategy_key, strategy_key)
        console.print(f"  [bold cyan]{strategy_key}[/] · [dim]{mname}[/]")

        t = Table(box=box.MINIMAL, show_header=True,
                  header_style="dim", show_lines=False)
        t.add_column("Professione O*NET", width=38)
        for r in res_list:
            t.add_column(r["title"][:14], justify="right", width=10)

        # Nota: il dettaglio per O*NET richiede gli score completi
        # che non sono salvati nella struttura corrente.
        # Mostriamo solo il best match per job.
        idx_map = {r["id"]: r for r in res_list}
        for j in onto_jobs:
            row_vals = [j["label"][:37]]
            for r in res_list:
                m   = r.get("match")
                raw = r.get("score_raw", 0.0)
                if m and m["job"]["uri"] == j["uri"]:
                    row_vals.append(f"[bold green]{m['score']:.3f}[/]")
                elif not m and r.get("score_raw", 0) > 0:
                    # Se questo è il candidato con score più alto (anche sotto soglia)
                    row_vals.append(f"[dim]{raw:.3f}[/]")
                else:
                    row_vals.append("[dim]—[/]")
            t.add_row(*row_vals)

        console.print(t)
        console.print()


def print_summary_multi(s1_res: list[dict],
                        nlp_results: dict[str, list[dict]],
                        model_names: dict[str, str]):
    n_total = len(s1_res)
    n_s1    = sum(1 for r in s1_res if r.get("match_s1"))

    lines = [f"  Job totali nel DB  : [bold]{n_total}[/]",
             f"  Match da S1        : [bold {'green' if n_s1 else 'red'}]{n_s1}/{n_total}[/]  [dim](soglia {S1_THRESHOLD})[/]"]

    for key in ["S2_MPNET", "S3_FINETUNED", "S4_MiniLM"]:
        res = nlp_results.get(key)
        if not res:
            lines.append(f"  Match da {key[:2]}         : [dim]— modello non disponibile —[/]")
            continue
        thr = MPNET_THRESHOLD if key == "S2_MPNET" else MINILM_THRESHOLD
        n   = sum(1 for r in res if r.get("match"))
        mname = model_names.get(key, key).split("/")[-1]
        lines.append(
            f"  Match da {key[:2]}         : [bold {'green' if n else 'red'}]{n}/{n_total}[/]"
            f"  [dim](soglia {thr} · {mname})[/]"
        )

    lines += [
        "",
        "  [dim]Nota: BERT base (S2) usa embedding [CLS] raw — non fine-tuned per sentence",
        "  similarity. Score tendenzialmente più bassi di MiniLM ma confrontabili tra loro.[/]",
    ]

    console.print(Panel("\n".join(lines),
                        title="[bold white]Riepilogo",
                        border_style="cyan", padding=(1, 2)))


# ═════════════════════════════════════════════════════════════
# OUTPUT EXCEL
# ═════════════════════════════════════════════════════════════

# ── Stili riutilizzabili ──────────────────────────────────────
_HDR_FILL   = PatternFill("solid", fgColor="1E3264")   # blu scuro
_HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_BODY_FONT  = Font(name="Arial", size=10)
_WRAP       = Alignment(wrap_text=True, vertical="top")
_CENTER     = Alignment(horizontal="center", vertical="center")
_THIN       = Side(style="thin", color="CCCCCC")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FILL_S1    = PatternFill("solid", fgColor="FFF3CD")   # giallo tenue
_FILL_BERT  = PatternFill("solid", fgColor="D0E8FB")   # azzurro tenue
_FILL_BIO   = PatternFill("solid", fgColor="F3D0FB")   # lilla tenue
_FILL_MINI  = PatternFill("solid", fgColor="D0FBE0")   # verde tenue
_FILL_ALT   = PatternFill("solid", fgColor="F8F8F8")   # grigio alternato

_GREEN_FONT = Font(name="Arial", size=10, color="1D6A38", bold=True)
_RED_FONT   = Font(name="Arial", size=10, color="A32D2D")
_DIM_FONT   = Font(name="Arial", size=10, color="888888", italic=True)


def _set_header(ws, row: int, cols: list[tuple[int, str]]):
    """Scrive una riga di intestazione con stile."""
    for col, text in cols:
        c = ws.cell(row=row, column=col, value=text)
        c.font      = _HDR_FONT
        c.fill      = _HDR_FILL
        c.alignment = _CENTER
        c.border    = _BORDER


def _style_cell(cell, fill=None, font=None, wrap=False, align=None):
    if fill:  cell.fill      = fill
    if font:  cell.font      = font or _BODY_FONT
    if not font: cell.font   = _BODY_FONT
    cell.border    = _BORDER
    cell.alignment = align or (_WRAP if wrap else Alignment(vertical="top"))


def save_excel(db_jobs: list[dict],
               s1_res:  list[dict],
               nlp_results: dict[str, list[dict]],
               model_names: dict[str, str],
               onto_jobs: list[dict]) -> str:
    """
    Produce un file Excel con tre fogli:
      1. Matching Results  — una riga per job, colonne per ogni strategia
      2. Ontology Jobs     — le 13 professioni O*NET con dettagli
      3. Score Detail      — tutti gli score raw per ogni job × strategia
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = os.path.join(OUTPUT_DIR, "rientra_matching.xlsx")
    wb   = Workbook()

    # Indici per lookup veloce
    s1_idx  = {r["id"]: r for r in s1_res}
    nlp_idx = {key: {r["id"]: r for r in res}
               for key, res in nlp_results.items() if res}
    avail   = [k for k in ["S2_MPNET", "S3_FINETUNED", "S4_MiniLM"] if k in nlp_idx]

    # ── FOGLIO 1: Matching Results ────────────────────────────
    ws1 = wb.active
    ws1.title = "Matching Results"
    ws1.freeze_panes = "A2"

    # Costruisce header dinamico
    base_cols = [
        (1, "ID"),
        (2, "Job Title (DB)"),
        (3, "Description (DB)"),
        (4, "IRI individuo"),
        # S1
        (5,  "S1 · Match O*NET"),
        (6,  "S1 · Score"),
        (7,  "S1 · Via label"),
        (8,  "S1 · O*NET URI"),
    ]
    nlp_start = 9
    nlp_col_map: dict[str, dict] = {}  # key → {match, score, uri}
    col = nlp_start
    for key in avail:
        short = model_names.get(key, key).split("/")[-1]
        nlp_col_map[key] = {"match": col, "score": col+1, "uri": col+2}
        base_cols += [
            (col,   f"{key[:2]} · {short} Match"),
            (col+1, f"{key[:2]} · Score"),
            (col+2, f"{key[:2]} · O*NET URI"),
        ]
        col += 3

    verdict_col = col
    base_cols.append((verdict_col, "Verdetto"))

    _set_header(ws1, 1, base_cols)

    # Larghezze colonne foglio 1
    col_widths_1 = {1:8, 2:28, 3:55, 4:35, 5:32, 6:10, 7:28, 8:55}
    for key in avail:
        cm = nlp_col_map[key]
        col_widths_1[cm["match"]] = 32
        col_widths_1[cm["score"]] = 10
        col_widths_1[cm["uri"]]   = 55
    col_widths_1[verdict_col] = 12
    for c, w in col_widths_1.items():
        ws1.column_dimensions[get_column_letter(c)].width = w

    for r_idx, row in enumerate(db_jobs, start=2):
        rid  = row["id"]
        alt  = (r_idx % 2 == 0)
        fill = _FILL_ALT if alt else None

        def wc(col, val, cfill=None, cfont=None, wrap=True):
            c = ws1.cell(row=r_idx, column=col, value=val)
            _style_cell(c, fill=cfill or fill, font=cfont, wrap=wrap)

        wc(1, rid,           wrap=False)
        wc(2, row["title"])
        wc(3, row.get("description",""))
        wc(4, f"{RIENTRA_BASE}ExtJob_{rid}")

        # S1
        m1 = s1_idx.get(rid, {}).get("match_s1")
        if m1:
            wc(5, m1["job"]["label"],   cfill=_FILL_S1)
            wc(6, round(m1["score"],4), cfill=_FILL_S1, wrap=False)
            wc(7, m1["via"],            cfill=_FILL_S1)
            wc(8, m1["job"]["uri"],     cfill=_FILL_S1)
        else:
            for c in [5,6,7,8]:
                cell = ws1.cell(row=r_idx, column=c, value="—")
                _style_cell(cell, fill=fill, font=_DIM_FONT, wrap=False)

        # NLP strategies
        uri_set = set()
        for key in avail:
            cm    = nlp_col_map[key]
            r2    = nlp_idx[key].get(rid, {})
            m2    = r2.get("match")
            raw   = r2.get("score_raw", 0.0)
            fills = {"S2_MPNET":      _FILL_BERT,
                     "S3_FINETUNED": _FILL_BIO,   # riuso colore lilla per fine-tuned
                     "S4_MiniLM":    _FILL_MINI}
            sf = fills.get(key)
            if m2:
                wc(cm["match"], m2["job"]["label"],    cfill=sf)
                sc_cell = ws1.cell(row=r_idx, column=cm["score"],
                                   value=round(m2["score"],4))
                _style_cell(sc_cell, fill=sf, wrap=False)
                sc_cell.font = _GREEN_FONT if m2["score"] >= 0.70 else _BODY_FONT
                wc(cm["uri"],   m2["job"]["uri"],       cfill=sf)
                uri_set.add(m2["job"]["uri"])
            else:
                no_cell = ws1.cell(row=r_idx, column=cm["match"],
                                   value=f"sotto soglia (raw {raw:.3f})")
                _style_cell(no_cell, fill=fill, font=_DIM_FONT)
                ws1.cell(row=r_idx, column=cm["score"],
                         value=round(raw,4)).font = _RED_FONT
                _style_cell(ws1.cell(row=r_idx, column=cm["score"]),
                            fill=fill, wrap=False)
                ws1.cell(row=r_idx, column=cm["uri"], value="—")
                _style_cell(ws1.cell(row=r_idx, column=cm["uri"]),
                            fill=fill, font=_DIM_FONT)

        # Verdetto
        if len(uri_set) == 1 and len(uri_set) == len(avail):
            verd, vfont = "Tutti concordano", Font(name="Arial",size=10,
                                                   color="1D6A38", bold=True)
        elif len(uri_set) == 1 and uri_set:
            verd, vfont = "Accordo parziale", Font(name="Arial",size=10,
                                                   color="85610A", bold=True)
        elif len(uri_set) > 1:
            verd, vfont = "Divergono", Font(name="Arial",size=10,
                                            color="A32D2D", bold=True)
        elif not uri_set and m1:
            verd, vfont = "Solo S1", Font(name="Arial",size=10,
                                          color="5C2D91", bold=True)
        else:
            verd, vfont = "Nessun match", Font(name="Arial",size=10,
                                               color="888888", italic=True)
        vc = ws1.cell(row=r_idx, column=verdict_col, value=verd)
        _style_cell(vc, fill=fill, font=vfont, wrap=False)

    ws1.row_dimensions[1].height = 30
    for r in range(2, len(db_jobs)+2):
        ws1.row_dimensions[r].height = 60

    # ── FOGLIO 2: Ontology Jobs ───────────────────────────────
    ws2 = wb.create_sheet("Ontology Jobs")
    ws2.freeze_panes = "A2"
    _set_header(ws2, 1, [
        (1,"#"),(2,"URI locale"),(3,"Label O*NET"),
        (4,"Net job titles (estratto)"),(5,"Descrizione O*NET"),
    ])
    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 48
    ws2.column_dimensions["C"].width = 36
    ws2.column_dimensions["D"].width = 50
    ws2.column_dimensions["E"].width = 65

    for i, j in enumerate(onto_jobs, 1):
        alt  = (i % 2 == 0)
        fill = _FILL_ALT if alt else None
        for col, val, wrap in [
            (1, i,               False),
            (2, j["local_name"], True),
            (3, j["label"],      True),
            (4, j["titles"][:300] if j["titles"] else "—", True),
            (5, j["description"][:500] if j["description"] else "—", True),
        ]:
            c = ws2.cell(row=i+1, column=col, value=val)
            _style_cell(c, fill=fill, wrap=wrap)
        ws2.row_dimensions[i+1].height = 60
    ws2.row_dimensions[1].height = 30

    # ── FOGLIO 3: Score Detail ────────────────────────────────
    ws3 = wb.create_sheet("Score Detail")
    ws3.freeze_panes = "C2"

    # Header: ID | Job Title | poi una colonna per ogni modello×professione
    hdr = [(1,"ID"),(2,"Job Title (DB)")]
    col = 3
    model_col_ranges: dict[str, tuple[int,int]] = {}
    for key in avail:
        start = col
        short = model_names.get(key, key).split("/")[-1]
        for j in onto_jobs:
            hdr.append((col, f"{short}\n{j['label'][:22]}"))
            col += 1
        model_col_ranges[key] = (start, col-1)
    _set_header(ws3, 1, hdr)
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 28
    for c in range(3, col):
        ws3.column_dimensions[get_column_letter(c)].width = 14

    # Righe dati con score raw per ogni combinazione
    for r_idx, row in enumerate(db_jobs, start=2):
        rid  = row["id"]
        ws3.cell(row=r_idx, column=1, value=rid)
        ws3.cell(row=r_idx, column=2, value=row["title"])
        for c in [1,2]:
            _style_cell(ws3.cell(row=r_idx, column=c),
                        fill=_FILL_ALT if r_idx%2==0 else None)

        for key in avail:
            r2       = nlp_idx[key].get(rid, {})
            m2       = r2.get("match")
            score_raw = r2.get("score_raw", 0.0)
            start_c, _ = model_col_ranges[key]

            # Per ora salviamo lo score best-match nella colonna della
            # professione trovata, e lasceremo vuote le altre.
            # (score completi richiederebbero di salvare l'intera matrice
            #  nel run_bert_matching — estensione futura)
            best_uri = m2["job"]["uri"] if m2 else None
            for c_off, j in enumerate(onto_jobs):
                c_abs = start_c + c_off
                is_best = (j["uri"] == best_uri)
                val = round(score_raw, 4) if is_best else None
                cell = ws3.cell(row=r_idx, column=c_abs, value=val)
                if is_best and val is not None:
                    fills = {"S2_MPNET": _FILL_BERT,
                             "S3_FINETUNED": _FILL_BIO,
                             "S4_MiniLM": _FILL_MINI}
                    _style_cell(cell, fill=fills.get(key),
                                font=_GREEN_FONT if score_raw >= 0.70
                                     else _BODY_FONT, wrap=False)
                else:
                    _style_cell(cell,
                                fill=_FILL_ALT if r_idx%2==0 else None,
                                wrap=False)

        ws3.row_dimensions[r_idx].height = 20
    ws3.row_dimensions[1].height = 45

    wb.save(path)
    return path


# ═════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Rientra@ Step 2 — Matching")
    parser.add_argument("--skip-training", action="store_true",
                        help="Salta il fine-tuning e usa il modello già salvato")
    args = parser.parse_args()

    console.print()
    console.print(Panel(
        f"[bold white]Rientra@[/] — Step 2: RDB2RDF + String Matching + NLP\n"
        f"[dim]Modelli: MPNet · MPNet fine-tuned · MiniLM  |  "
        f"Tutti sentence-transformers nativi  |  Metrica: cosine similarity[/]\n"
        f"[dim]{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}[/]",
        border_style="blue", padding=(0, 2),
    ))
    console.print()

    # ── [0] Setup ─────────────────────────────────────────────
    console.print(Rule("[bold cyan]0 · Setup[/]", style="cyan"))
    console.print()

    if not test_connection():
        sys.exit(1)

    with console.status(f"  Caricamento ontologia [cyan]{ONTOLOGY_FILE}[/]...",
                        spinner="dots"):
        onto_graph = load_ontology(ONTOLOGY_FILE)
    console.print(f"  [bold green]✓[/] Ontologia: [bold]{len(onto_graph)}[/] triple")

    onto_jobs = extract_ontology_jobs(onto_graph)
    console.print(f"  [bold green]✓[/] Professioni O*NET estratte: [bold]{len(onto_jobs)}[/]")
    console.print()
    print_ontology_table(onto_jobs)
    console.print()

    # ── [1] Lettura DB ────────────────────────────────────────
    console.print(Rule("[bold magenta]1 · Lettura DB esterno[/]", style="magenta"))
    console.print()
    db_jobs = fetch_all(
        "SELECT id, title, description FROM ext_job ORDER BY id")
    console.print(
        f"  [bold green]✓[/] [bold]{len(db_jobs)}[/] lavori da [cyan]ext_job[/]")
    console.print()
    print_db_table(db_jobs)
    console.print()

    # ── [2] Strategia 1 — String Matching ─────────────────────
    console.print(Rule("[bold yellow]2 · Strategia 1 — String Matching (pre-processing esteso)[/]",
                       style="yellow"))
    console.print(
        "  [dim]Jaccard (0.5) + Overlap (0.3) + Levenshtein (0.2)[/]\n"
        "  [dim]Pre-processing: lowercase · simboli → spazio · "
        "espansione abbreviazioni · rimozione stopwords[/]")
    console.print()
    s1_res = run_string_matching(db_jobs, onto_jobs)
    console.print(f"  [bold green]✓[/] Completato\n")

    # ── [3] Strategia 2 — MPNet sentence-transformer ──────────
    console.print(Rule("[bold blue]3 · Strategia 2 — all-mpnet-base-v2[/]",
                       style="blue"))
    console.print(
        f"  [dim]Modello: {MODELS['S2_MPNET']}[/]\n"
        f"  [dim]Sentence-transformer nativo · 768 dim · Soglia: {MPNET_THRESHOLD}[/]\n"
        f"  [dim]Ottimizzato per testi lunghi · nessun problema di anisotropy[/]")
    console.print()
    s2_res, s2_name = run_st_matching(
        db_jobs, onto_jobs,
        model_name_or_path=MODELS["S2_MPNET"],
        strategy_key="S2_MPNET",
        threshold=MPNET_THRESHOLD,
    )
    if s2_res:
        console.print(f"  [bold green]✓[/] Completato\n")
    else:
        console.print(
            f"  [yellow]⚠[/] Saltato — installa [cyan]sentence-transformers[/]\n")

    # ── [4] Strategia 3 — MPNet fine-tuned ───────────────────
    console.print(Rule("[bold magenta]4 · Strategia 3 — MPNet fine-tuned (Rientra@)[/]",
                       style="magenta"))
    console.print(
        f"  [dim]Base: {MODELS['S2_MPNET']}[/]\n"
        f"  [dim]Training: {TRAINING_FILE}  |  Soglia: {FINETUNED_THRESHOLD}[/]\n"
        f"  [dim]Loss: CosineSimilarityLoss  |  "
        f"Epoche: {FT_EPOCHS}  |  Batch: {FT_BATCH_SIZE}[/]")
    console.print()

    s3_res, s3_name = None, None
    model_already_saved = os.path.isdir(FINETUNED_DIR) and any(
        os.scandir(FINETUNED_DIR))

    if args.skip_training and model_already_saved:
        console.print(
            f"  [dim]--skip-training attivo: riuso modello in [cyan]{FINETUNED_DIR}[/][/]")
        s3_res, s3_name = run_finetuned_matching(
            db_jobs, onto_jobs, FINETUNED_DIR, FINETUNED_THRESHOLD)

    elif not os.path.exists(TRAINING_FILE):
        console.print(
            f"  [yellow]⚠[/] File di training [cyan]{TRAINING_FILE}[/] non trovato.\n"
            f"  Genera il file con: [cyan]python gen_training_data.py[/]\n"
            f"  Poi valida le coppie nel foglio 'Training Pairs' "
            f"(colonna Validated = SI/NO).\n")

    else:
        # Carica le coppie di training
        console.print("  [dim]Caricamento coppie di training...[/]")
        training_pairs = load_training_pairs(TRAINING_FILE)

        if model_already_saved and not args.skip_training:
            console.print(
                f"  [dim]Modello già presente in [cyan]{FINETUNED_DIR}[/]. "
                f"Usa --skip-training per riutilizzarlo.[/]")
            console.print(
                "  [dim]Avvio nuovo training (sovrascrive il modello precedente)...[/]")

        # Fine-tuning
        console.print()
        ft_path = finetune_mpnet(training_pairs, FINETUNED_DIR)

        if ft_path:
            console.print()
            s3_res, s3_name = run_finetuned_matching(
                db_jobs, onto_jobs, ft_path, FINETUNED_THRESHOLD)
            if s3_res:
                console.print(f"  [bold green]✓[/] Completato\n")
        else:
            console.print(
                f"  [yellow]⚠[/] Fine-tuning non riuscito — installa "
                f"[cyan]sentence-transformers torch[/]\n")

    # ── [5] Strategia 4 — MiniLM ──────────────────────────────
    console.print(Rule("[bold green]5 · Strategia 4 — MiniLM (riferimento)[/]",
                       style="green"))
    console.print(
        f"  [dim]Modello: {MODELS['S4_MiniLM']}  |  "
        f"Fine-tuned per sentence similarity  |  Soglia: {MINILM_THRESHOLD}[/]")
    console.print()
    s4_res, s4_name = run_minilm_matching(
        db_jobs, onto_jobs, threshold=MINILM_THRESHOLD)
    if s4_res:
        console.print(f"  [bold green]✓[/] Completato\n")
    else:
        console.print(
            f"  [yellow]⚠[/] Saltato — installa [cyan]sentence-transformers[/]\n")

    # Raccoglie risultati disponibili
    nlp_results: dict[str, list[dict]] = {}
    model_names: dict[str, str]        = {}
    if s2_res:
        nlp_results["S2_MPNET"]     = s2_res
        model_names["S2_MPNET"]     = s2_name
    if s3_res:
        nlp_results["S3_FINETUNED"] = s3_res
        model_names["S3_FINETUNED"] = s3_name
    if s4_res:
        nlp_results["S4_MiniLM"]    = s4_res
        model_names["S4_MiniLM"]    = s4_name

    # ── [6] Confronto ─────────────────────────────────────────
    console.print(Rule("[bold white]6 · Confronto strategie[/]", style="white"))
    console.print()
    if nlp_results:
        print_full_comparison(s1_res, nlp_results, model_names)
        print_summary_multi(s1_res, nlp_results, model_names)
    else:
        console.print(
            "  [yellow]Nessun modello NLP disponibile — solo S1[/]")
    console.print()

    # ── [7] Output Excel ──────────────────────────────────────
    console.print(Rule("[bold blue]7 · Produzione file Excel[/]", style="blue"))
    console.print()
    with console.status("  Costruzione workbook...", spinner="dots"):
        xlsx_path = save_excel(db_jobs, s1_res, nlp_results,
                               model_names, onto_jobs)

    size_kb = os.path.getsize(xlsx_path) // 1024
    t = Table(box=box.SIMPLE, show_header=False, padding=(0, 1))
    t.add_column(style="dim")
    t.add_column(style="cyan")
    t.add_column(style="dim", justify="right")
    t.add_row("Excel",    xlsx_path,           f"{size_kb:,} KB")
    t.add_row("Foglio 1", "Matching Results",
              f"{len(db_jobs)} righe · tutte le strategie")
    t.add_row("Foglio 2", "Ontology Jobs",
              f"{len(onto_jobs)} professioni O*NET")
    t.add_row("Foglio 3", "Score Detail",
              "score per job × modello")
    console.print(t)
    console.print()

    # ── Fine ──────────────────────────────────────────────────
    ft_note = (f"Modello fine-tuned salvato in: [cyan]{FINETUNED_DIR}[/]\n"
               if s3_res else
               "Fine-tuned non disponibile in questa esecuzione.\n")
    console.print(Panel(
        f"[bold green]Step 2 completato.[/]\n\n"
        f"Output → [cyan]{xlsx_path}[/]\n"
        f"{ft_note}\n"
        f"[dim]Modelli usati: "
        f"{', '.join(model_names.values()) or 'solo S1'}[/]\n"
        f"[dim]Prossimo step: Strategia 3 -- ESCO come ponte ontologico[/]",
        border_style="green", padding=(0, 2),
    ))
    console.print()


if __name__ == "__main__":
    main()