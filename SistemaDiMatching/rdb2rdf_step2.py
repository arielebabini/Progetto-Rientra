"""
rdb2rdf_step2.py
================
Step 2 del progetto Rientra@: RDB2RDF con matching verso l'ontologia principale.

Dataset in ingresso:
  - Tabella PostgreSQL  : ext_job  (ext01…ext09)
    → creare con: psql -U postgres -d rientra_db -f rientra_ext_jobs.sql
  - Ontologia           : Rientra.rdf  (namespace http://www.stiima.cnr.it/JobList#)

Strategie di matching:
  S1 — String Matching   : Jaccard + Overlap + Levenshtein (baseline lessicale)
  S2 — BERT base         : bert-base-uncased, embedding [CLS], cosine similarity
  S3 — BioBERT           : dmis-lab/biobert-base-cased-v1.2, embedding [CLS], cosine similarity
  S4 — MiniLM            : all-MiniLM-L6-v2 via sentence-transformers (riferimento)

Nota sui modelli BERT "raw":
  BERT base e BioBERT non sono modelli sentence-level nativi come MiniLM.
  Per ricavare un embedding di frase si usa il token [CLS] dell'ultimo hidden
  state — tecnica standard per task di classificazione e similarità con BERT.
  BioBERT è BERT fine-tuned su letteratura biomedica (PubMed + PMC): non è
  il dominio ideale per job matching, ma è incluso per confronto metodologico.

Requisiti:
    pip install psycopg2-binary rdflib rich transformers torch scikit-learn
    pip install sentence-transformers   # per S4 (MiniLM)

Uso:
    python rdb2rdf_step2.py
"""

import os
import sys
import numpy as np
from datetime import datetime

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich import box
from rich.text import Text
from rich.rule import Rule

console = Console()

# ─────────────────────────────────────────────────────────────
# CONFIGURAZIONE
# ─────────────────────────────────────────────────────────────

DB_CONFIG = {
    "host":     "localhost",
    "port":     5432,
    "dbname":   "rientra_db",
    "user":     "postgres",
    "password": "postgres",
}

ONTOLOGY_FILE = "Rientra.rdf"
OUTPUT_DIR    = "output"

JOBLIST_BASE = "http://www.stiima.cnr.it/JobList#"
RIENTRA_BASE = "https://www.stiima.cnr.it/rientra#"

# Soglie cosine similarity per i modelli NLP
# (i modelli BERT raw tendono a score più bassi di MiniLM per sentence similarity)
S1_THRESHOLD    = 0.12
BERT_THRESHOLD  = 0.70   # BERT base e BioBERT: soglia più alta perché [CLS] è meno discriminante
MINILM_THRESHOLD = 0.50  # MiniLM: fine-tuned per sentence similarity, score più affidabili

# Modelli da usare
MODELS = {
    "S2_BERT":    "bert-base-uncased",
    "S3_BioBERT": "dmis-lab/biobert-base-cased-v1.2",
    "S4_MiniLM":  "all-MiniLM-L6-v2",   # via sentence-transformers
}


# ─────────────────────────────────────────────────────────────
# HELPERS VISIVI
# ─────────────────────────────────────────────────────────────

def score_color(score: float) -> str:
    if score >= 0.85: return "bold green"
    if score >= 0.70: return "green"
    if score >= 0.50: return "yellow"
    if score >= 0.30: return "orange3"
    return "red"

def score_bar(score: float, width: int = 8) -> str:
    filled = round(score * width)
    return "█" * filled + "░" * (width - filled)


# ═════════════════════════════════════════════════════════════
# UTILITÀ DB
# ═════════════════════════════════════════════════════════════

def get_connection():
    return psycopg2.connect(**DB_CONFIG)

def fetch_all(sql: str, params=None) -> list[dict]:
    with get_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            return [dict(r) for r in cur.fetchall()]

def test_connection() -> bool:
    try:
        with get_connection() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT version();")
                v = cur.fetchone()[0]
        console.print(f"  [bold green]✓[/] PostgreSQL [dim]{v[:60]}...[/]")
        return True
    except Exception as e:
        console.print(f"  [bold red]✗[/] Errore connessione: [red]{e}[/]")
        return False


# ═════════════════════════════════════════════════════════════
# ESTRAZIONE VOCABOLARIO DALL'ONTOLOGIA
# ═════════════════════════════════════════════════════════════

def load_ontology(path: str):
    """Carica il file RDF e restituisce un oggetto Graph rdflib."""
    from rdflib import Graph
    g = Graph()
    g.parse(path, format="xml")
    return g

def extract_ontology_jobs(onto_graph) -> list[dict]:
    from rdflib import RDF, OWL
    from rdflib.namespace import RDFS
    jobs = []
    for s, _, _ in onto_graph.triples((None, RDF.type, OWL.Class)):
        if not str(s).startswith(JOBLIST_BASE):
            continue
        local = str(s).split("#")[1]
        if local in ("Job", "Job_Descriptor"):
            continue

        label = str(onto_graph.value(s, RDFS.label) or "").strip()
        titles, desc = "", ""
        for pred, obj in onto_graph.predicate_objects(s):
            ps = str(pred)
            if "Net_job_titles" in ps or "job_titles" in ps:
                titles = str(obj)
            if "Net_short_description" in ps or "short_description" in ps:
                desc = str(obj)

        if titles and ":" in titles:
            titles = titles.split(":", 1)[1].strip()

        display = label or local.replace("_", " ")
        jobs.append({
            "uri":         str(s),
            "local_name":  local,
            "label":       display,
            "titles":      titles,
            "description": desc,
        })

    jobs.sort(key=lambda x: x["label"])
    return jobs

def _build_onto_text(j: dict) -> str:
    """Testo composito per l'ontologia: label + titoli alternativi + descrizione."""
    parts = [j["label"]]
    if j["titles"]:
        parts.append(j["titles"])
    if j["description"]:
        parts.append(j["description"])
    return " ".join(parts)


# ═════════════════════════════════════════════════════════════
# STRATEGIA 1 — STRING MATCHING
# ═════════════════════════════════════════════════════════════

def _norm(s: str) -> str:
    return s.lower().strip()

def _jaccard(a: str, b: str) -> float:
    a, b = set(_norm(a).split()), set(_norm(b).split())
    return len(a & b) / len(a | b) if (a | b) else 1.0

def _overlap(a: str, b: str) -> float:
    a, b = set(_norm(a).split()), set(_norm(b).split())
    return len(a & b) / min(len(a), len(b)) if (a and b) else 0.0

def _levenshtein(s1: str, s2: str) -> float:
    s1, s2 = _norm(s1), _norm(s2)
    m, n = len(s1), len(s2)
    if not m and not n:
        return 1.0
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev, dp[0] = dp[:], i
        for j in range(1, n + 1):
            cost = 0 if s1[i-1] == s2[j-1] else 1
            dp[j] = min(prev[j] + 1, dp[j-1] + 1, prev[j-1] + cost)
    return 1.0 - dp[n] / max(m, n)

def _combined(a: str, b: str) -> float:
    return 0.5 * _jaccard(a, b) + 0.3 * _overlap(a, b) + 0.2 * _levenshtein(a, b)

def s1_match_one(title: str, onto_jobs: list[dict]) -> dict | None:
    best, bj, bv = 0.0, None, None
    for j in onto_jobs:
        cands = [j["label"]]
        if j["titles"]:
            cands += [t.strip() for t in j["titles"].split(",") if t.strip()]
        for c in cands:
            sc = _combined(title, c)
            if sc > best:
                best, bj, bv = sc, j, c
    if best >= S1_THRESHOLD:
        return {"job": bj, "score": best, "via": bv}
    return None

def run_string_matching(db_jobs: list[dict], onto_jobs: list[dict]) -> list[dict]:
    results = []
    with Progress(SpinnerColumn(), TextColumn("{task.description}"),
                  BarColumn(), TaskProgressColumn(),
                  console=console, transient=True) as p:
        task = p.add_task("  Calcolo...", total=len(db_jobs))
        for r in db_jobs:
            results.append({"id": r["id"], "title": r["title"],
                            "match_s1": s1_match_one(r["title"], onto_jobs)})
            p.advance(task)
    return results


# ═════════════════════════════════════════════════════════════
# COSINE SIMILARITY — funzione centrale condivisa da S2 e S3
# ═════════════════════════════════════════════════════════════

def cosine_similarity_matrix(query_emb: np.ndarray,
                              corpus_embs: np.ndarray) -> np.ndarray:
    """
    Calcola la cosine similarity tra un vettore query e una matrice di vettori corpus.

    Formula:
        cos(θ) = (A · B) / (‖A‖ · ‖B‖)

    dove A è il vettore del job nel DB e B è il vettore di una professione O*NET.

    Il risultato è compreso tra -1 e 1:
      1.0  → testi identici nel significato
      0.0  → testi ortogonali (nessuna relazione semantica)
     -1.0  → testi opposti (raro con embedding linguistici)

    Per embedding BERT non negativi il range effettivo è [0, 1].
    """
    query_norm  = query_emb / (np.linalg.norm(query_emb) + 1e-9)
    corpus_norm = corpus_embs / (np.linalg.norm(corpus_embs, axis=1, keepdims=True) + 1e-9)
    return corpus_norm.dot(query_norm)


# ═════════════════════════════════════════════════════════════
# EMBEDDING CON BERT RAW (BERT base e BioBERT)
# ═════════════════════════════════════════════════════════════

def get_bert_embedding(text: str, tokenizer, model,
                       max_length: int = 512) -> np.ndarray:
    """
    Calcola l'embedding di una frase con un modello BERT raw.

    Strategia: estrae il vettore del token [CLS] dall'ultimo hidden state.
    Il token [CLS] (Classification) è il primo token di ogni sequenza BERT
    ed è stato progettato per catturare una rappresentazione aggregata
    dell'intera sequenza — per questo è la scelta standard per embedding
    di frasi con BERT non fine-tuned su sentence similarity.

    Alternativa possibile: mean pooling su tutti i token non-padding,
    che spesso produce embedding leggermente più stabili ma richiede
    la gestione della attention mask.
    """
    import torch
    inputs = tokenizer(
        text,
        return_tensors="pt",
        truncation=True,
        max_length=max_length,
        padding=True,
    )
    with torch.no_grad():
        outputs = model(**inputs)
    # outputs.last_hidden_state: shape [batch, seq_len, hidden_size]
    # [:, 0, :] → token [CLS], primo token della sequenza
    cls_embedding = outputs.last_hidden_state[:, 0, :].squeeze().numpy()
    return cls_embedding


def run_bert_matching(db_jobs: list[dict],
                      onto_jobs: list[dict],
                      model_name: str,
                      strategy_key: str,
                      threshold: float) -> tuple[list[dict], str]:
    """
    Esegue il matching con un modello BERT raw (bert-base o biobert).
    Restituisce (risultati, nome_modello) oppure (None, None) se il modello
    non è disponibile.
    """
    try:
        from transformers import AutoTokenizer, AutoModel

        with console.status(f"  Caricamento [cyan]{model_name}[/]...", spinner="dots"):
            tokenizer = AutoTokenizer.from_pretrained(model_name)
            model     = AutoModel.from_pretrained(model_name)
            model.eval()
        console.print(f"  [bold green]✓[/] Modello caricato: [cyan]{model_name}[/]")

        onto_texts = [_build_onto_text(j) for j in onto_jobs]
        db_texts   = [f"{r['title']}. {r.get('description', '')}" for r in db_jobs]

        # Pre-calcola embedding O*NET (una volta sola)
        with Progress(SpinnerColumn(), TextColumn("{task.description}"),
                      BarColumn(), TaskProgressColumn(),
                      console=console, transient=True) as p:
            t1 = p.add_task("  Encoding O*NET...", total=len(onto_texts))
            onto_embs = []
            for txt in onto_texts:
                onto_embs.append(get_bert_embedding(txt, tokenizer, model))
                p.advance(t1)
        onto_embs = np.vstack(onto_embs)   # shape: [n_onto, hidden_size]

        # Calcola embedding per ogni job del DB e cosine similarity
        results = []
        with Progress(SpinnerColumn(), TextColumn("{task.description}"),
                      BarColumn(), TaskProgressColumn(),
                      console=console, transient=True) as p:
            t2 = p.add_task("  Cosine similarity...", total=len(db_jobs))
            for i, row in enumerate(db_jobs):
                q_emb = get_bert_embedding(db_texts[i], tokenizer, model)
                sims  = cosine_similarity_matrix(q_emb, onto_embs)
                idx   = int(np.argmax(sims))
                sc    = float(sims[idx])
                m = {"job": onto_jobs[idx], "score": sc} if sc >= threshold else None
                results.append({
                    "id":        row["id"],
                    "title":     row["title"],
                    "match":     m,
                    "score_raw": sc,
                    "strategy":  strategy_key,
                })
                p.advance(t2)

        return results, model_name

    except Exception as e:
        console.print(f"  [yellow]⚠[/] {model_name} non disponibile: [dim]{e}[/]")
        return None, None


# ═════════════════════════════════════════════════════════════
# EMBEDDING CON SENTENCE-TRANSFORMERS (MiniLM — S4)
# ═════════════════════════════════════════════════════════════

def run_minilm_matching(db_jobs: list[dict],
                        onto_jobs: list[dict],
                        threshold: float) -> tuple[list[dict], str]:
    """
    Esegue il matching con sentence-transformers/all-MiniLM-L6-v2.
    A differenza di BERT raw, MiniLM è fine-tuned per sentence similarity:
    il suo .encode() produce già embedding ottimizzati per cosine similarity,
    senza bisogno di estrarre manualmente il token [CLS].
    La cosine similarity viene comunque calcolata esplicitamente con la
    nostra funzione cosine_similarity_matrix per uniformità metodologica.
    """
    try:
        from sentence_transformers import SentenceTransformer

        model_name = MODELS["S4_MiniLM"]
        with console.status(f"  Caricamento [cyan]{model_name}[/]...", spinner="dots"):
            model = SentenceTransformer(model_name)
        console.print(f"  [bold green]✓[/] Modello caricato: [cyan]{model_name}[/]")

        onto_texts = [_build_onto_text(j) for j in onto_jobs]
        db_texts   = [f"{r['title']}. {r.get('description', '')}" for r in db_jobs]

        with Progress(SpinnerColumn(), TextColumn("{task.description}"),
                      BarColumn(), TaskProgressColumn(),
                      console=console, transient=True) as p:
            t1 = p.add_task("  Encoding O*NET...", total=1)
            onto_embs = model.encode(onto_texts, convert_to_numpy=True)
            p.advance(t1)

            results = []
            t2 = p.add_task("  Cosine similarity...", total=len(db_jobs))
            for i, row in enumerate(db_jobs):
                q_emb = model.encode(db_texts[i], convert_to_numpy=True)
                # Cosine similarity calcolata esplicitamente (stessa formula di S2/S3)
                sims  = cosine_similarity_matrix(q_emb, onto_embs)
                idx   = int(np.argmax(sims))
                sc    = float(sims[idx])
                m = {"job": onto_jobs[idx], "score": sc} if sc >= threshold else None
                results.append({
                    "id":        row["id"],
                    "title":     row["title"],
                    "match":     m,
                    "score_raw": sc,
                    "strategy":  "S4_MiniLM",
                })
                p.advance(t2)

        return results, model_name

    except Exception as e:
        console.print(f"  [yellow]⚠[/] MiniLM non disponibile: [dim]{e}[/]")
        return None, None


# ═════════════════════════════════════════════════════════════
# OUTPUT RICH
# ═════════════════════════════════════════════════════════════

def print_ontology_table(onto_jobs: list[dict]):
    t = Table(title="Professioni O*NET nell'ontologia",
              box=box.ROUNDED, header_style="bold cyan",
              title_style="bold white", show_lines=False)
    t.add_column("#",           style="dim",       width=3,  justify="right")
    t.add_column("URI locale",  style="cyan",       width=46)
    t.add_column("Label O*NET", style="bold white", width=38)
    t.add_column("Desc.",       justify="center",   width=6)
    for i, j in enumerate(onto_jobs, 1):
        t.add_row(str(i), j["local_name"], j["label"],
                  "[green]✓[/]" if j["description"] else "[red]✗[/]")
    console.print(t)


def print_db_table(db_jobs: list[dict]):
    t = Table(title="Lavori nel DB esterno (ext_job)",
              box=box.ROUNDED, header_style="bold magenta",
              title_style="bold white", show_lines=False)
    t.add_column("ID",    style="bold magenta", width=6)
    t.add_column("Titolo",                      width=30)
    t.add_column("Descrizione (estratto)",      width=62)
    for r in db_jobs:
        desc = (r.get("description") or "")[:60] + "…"
        t.add_row(r["id"], r["title"], f"[dim]{desc}[/]")
    console.print(t)


def _score_cell(score: float | None, raw: float = 0.0) -> tuple[Text, Text]:
    """Restituisce (cella_label_aggiuntiva, cella_score) per la tabella."""
    if score is not None:
        bar = Text()
        bar.append(f"{score:.3f}\n", style=score_color(score))
        bar.append(score_bar(score),  style=score_color(score))
        return Text(""), bar
    else:
        lbl = Text(f"(raw: {raw:.3f})", style="dim red")
        bar = Text()
        bar.append(f"{raw:.3f}\n", style="dim red")
        bar.append(score_bar(raw),  style="dim red")
        return lbl, bar


def print_full_comparison(s1_res: list[dict],
                          nlp_results: dict[str, list[dict]],
                          model_names: dict[str, str]):
    """
    Tabella unica con S1 + tutti i modelli NLP disponibili, una colonna per modello.
    nlp_results: { "S2_BERT": [...], "S3_BioBERT": [...], "S4_MiniLM": [...] }
    model_names: { "S2_BERT": "bert-base-uncased", ... }
    """
    # Indici per lookup rapido per ogni strategia
    idx = {key: {r["id"]: r for r in res}
           for key, res in nlp_results.items() if res}

    available_strategies = [k for k in ["S2_BERT", "S3_BioBERT", "S4_MiniLM"]
                            if k in idx]

    # Costruisce header dinamico
    title_parts = ["S1 String"] + [
        model_names.get(k, k).split("/")[-1][:18]
        for k in available_strategies
    ]
    t = Table(
        title="Confronto strategie di matching · cosine similarity",
        box=box.SIMPLE_HEAD,
        show_lines=True,
        header_style="bold white on dark_blue",
        title_style="bold white",
        expand=True,
    )
    t.add_column("ID",       style="bold", width=6,  justify="center")
    t.add_column("Titolo DB",              width=26)
    t.add_column("S1 — Match",             width=28)
    t.add_column("Score S1", justify="center", width=14)

    for k in available_strategies:
        short = model_names.get(k, k).split("/")[-1][:20]
        t.add_column(f"{k[:2]} — {short}", width=28)
        t.add_column(f"Score {k[:2]}", justify="center", width=14)

    t.add_column("Verdetto", justify="center", width=10)

    for r in s1_res:
        rid = r["id"]
        m1  = r.get("match_s1")

        # S1
        if m1:
            s1_cell = Text()
            s1_cell.append(m1["job"]["label"] + "\n")
            s1_cell.append(f"via: {m1['via'][:24]}", style="dim")
            s1_sc = Text()
            s1_sc.append(f"{m1['score']:.3f}\n", style=score_color(m1["score"]))
            s1_sc.append(score_bar(m1["score"]),  style=score_color(m1["score"]))
        else:
            s1_cell = Text("— nessun match —", style="dim")
            s1_sc   = Text("—", style="dim")

        row_cells = [rid, r["title"], s1_cell, s1_sc]

        # NLP strategies
        nlp_matches = []
        for k in available_strategies:
            r2   = idx[k].get(rid, {})
            m2   = r2.get("match")
            raw  = r2.get("score_raw", 0.0)
            if m2:
                cell = Text(m2["job"]["label"])
                sc   = Text()
                sc.append(f"{m2['score']:.3f}\n", style=score_color(m2["score"]))
                sc.append(score_bar(m2["score"]),  style=score_color(m2["score"]))
                nlp_matches.append(m2["job"]["uri"])
            else:
                cell = Text()
                cell.append("— sotto soglia —\n", style="dim")
                cell.append(f"(raw: {raw:.3f})", style="dim red")
                sc   = Text()
                sc.append(f"{raw:.3f}\n", style="dim red")
                sc.append(score_bar(raw),  style="dim red")
                nlp_matches.append(None)
            row_cells += [cell, sc]

        # Verdetto: consensus tra i modelli che hanno trovato un match
        valid = [u for u in nlp_matches if u]
        if not valid and not m1:
            verdetto = Text("✗ no", style="bold red")
        elif len(set(valid)) == 1 and len(valid) == len(available_strategies):
            verdetto = Text("✓ tutti", style="bold green")
        elif len(set(valid)) == 1 and valid:
            verdetto = Text("~ parz.", style="yellow")
        elif valid:
            verdetto = Text("≠ div.", style="orange3")
        else:
            verdetto = Text("S1 only", style="magenta")

        row_cells.append(verdetto)
        t.add_row(*row_cells)

    console.print(t)
    console.print(
        "  [bold green]✓ tutti[/] tutti i modelli NLP concordano  "
        "[yellow]~ parz.[/] accordo parziale  "
        "[orange3]≠ div.[/] modelli divergono  "
        "[magenta]S1 only[/] solo string match  "
        "[bold red]✗ no[/] nessun match\n"
    )


def print_cosine_detail(nlp_results: dict[str, list[dict]],
                        model_names: dict[str, str],
                        onto_jobs: list[dict]):
    """
    Tabella dettaglio cosine similarity: per ogni job DB mostra lo score
    con TUTTE le professioni O*NET, per ogni modello disponibile.
    Utile per capire la distribuzione degli score e validare le soglie.
    """
    console.print(Rule("[dim]Dettaglio score cosine per tutti i candidati O*NET[/]", style="dim"))
    console.print("[dim]  (i valori in verde sono quelli scelti come match migliore)[/]\n")

    available = [k for k in ["S2_BERT", "S3_BioBERT", "S4_MiniLM"]
                 if k in nlp_results and nlp_results[k]]

    for strategy_key in available:
        res_list = nlp_results[strategy_key]
        mname    = model_names.get(strategy_key, strategy_key)
        console.print(f"  [bold cyan]{strategy_key}[/] · [dim]{mname}[/]")

        t = Table(box=box.MINIMAL, show_header=True,
                  header_style="dim", show_lines=False)
        t.add_column("Professione O*NET", width=38)
        for r in res_list:
            t.add_column(r["title"][:14], justify="right", width=10)

        # Nota: il dettaglio per O*NET richiede gli score completi
        # che non sono salvati nella struttura corrente.
        # Mostriamo solo il best match per job.
        idx_map = {r["id"]: r for r in res_list}
        for j in onto_jobs:
            row_vals = [j["label"][:37]]
            for r in res_list:
                m   = r.get("match")
                raw = r.get("score_raw", 0.0)
                if m and m["job"]["uri"] == j["uri"]:
                    row_vals.append(f"[bold green]{m['score']:.3f}[/]")
                elif not m and r.get("score_raw", 0) > 0:
                    # Se questo è il candidato con score più alto (anche sotto soglia)
                    row_vals.append(f"[dim]{raw:.3f}[/]")
                else:
                    row_vals.append("[dim]—[/]")
            t.add_row(*row_vals)

        console.print(t)
        console.print()


def print_summary_multi(s1_res: list[dict],
                        nlp_results: dict[str, list[dict]],
                        model_names: dict[str, str]):
    n_total = len(s1_res)
    n_s1    = sum(1 for r in s1_res if r.get("match_s1"))

    lines = [f"  Job totali nel DB  : [bold]{n_total}[/]",
             f"  Match da S1        : [bold {'green' if n_s1 else 'red'}]{n_s1}/{n_total}[/]  [dim](soglia {S1_THRESHOLD})[/]"]

    for key in ["S2_BERT", "S3_BioBERT", "S4_MiniLM"]:
        res = nlp_results.get(key)
        if not res:
            lines.append(f"  Match da {key[:2]}         : [dim]— modello non disponibile —[/]")
            continue
        thr = BERT_THRESHOLD if key in ("S2_BERT", "S3_BioBERT") else MINILM_THRESHOLD
        n   = sum(1 for r in res if r.get("match"))
        mname = model_names.get(key, key).split("/")[-1]
        lines.append(
            f"  Match da {key[:2]}         : [bold {'green' if n else 'red'}]{n}/{n_total}[/]"
            f"  [dim](soglia {thr} · {mname})[/]"
        )

    lines += [
        "",
        "  [dim]Nota: BERT base e BioBERT usano embedding [CLS] — non fine-tuned per sentence",
        "  similarity. Score tendenzialmente più bassi di MiniLM ma confrontabili tra loro.[/]",
    ]

    console.print(Panel("\n".join(lines),
                        title="[bold white]Riepilogo",
                        border_style="cyan", padding=(1, 2)))


# ═════════════════════════════════════════════════════════════
# OUTPUT EXCEL
# ═════════════════════════════════════════════════════════════

# ── Stili riutilizzabili ──────────────────────────────────────
_HDR_FILL   = PatternFill("solid", fgColor="1E3264")   # blu scuro
_HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
_BODY_FONT  = Font(name="Arial", size=10)
_WRAP       = Alignment(wrap_text=True, vertical="top")
_CENTER     = Alignment(horizontal="center", vertical="center")
_THIN       = Side(style="thin", color="CCCCCC")
_BORDER     = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FILL_S1    = PatternFill("solid", fgColor="FFF3CD")   # giallo tenue
_FILL_BERT  = PatternFill("solid", fgColor="D0E8FB")   # azzurro tenue
_FILL_BIO   = PatternFill("solid", fgColor="F3D0FB")   # lilla tenue
_FILL_MINI  = PatternFill("solid", fgColor="D0FBE0")   # verde tenue
_FILL_ALT   = PatternFill("solid", fgColor="F8F8F8")   # grigio alternato

_GREEN_FONT = Font(name="Arial", size=10, color="1D6A38", bold=True)
_RED_FONT   = Font(name="Arial", size=10, color="A32D2D")
_DIM_FONT   = Font(name="Arial", size=10, color="888888", italic=True)


def _set_header(ws, row: int, cols: list[tuple[int, str]]):
    """Scrive una riga di intestazione con stile."""
    for col, text in cols:
        c = ws.cell(row=row, column=col, value=text)
        c.font      = _HDR_FONT
        c.fill      = _HDR_FILL
        c.alignment = _CENTER
        c.border    = _BORDER


def _style_cell(cell, fill=None, font=None, wrap=False, align=None):
    if fill:  cell.fill      = fill
    if font:  cell.font      = font or _BODY_FONT
    if not font: cell.font   = _BODY_FONT
    cell.border    = _BORDER
    cell.alignment = align or (_WRAP if wrap else Alignment(vertical="top"))


def save_excel(db_jobs: list[dict],
               s1_res:  list[dict],
               nlp_results: dict[str, list[dict]],
               model_names: dict[str, str],
               onto_jobs: list[dict]) -> str:
    """
    Produce un file Excel con tre fogli:
      1. Matching Results  — una riga per job, colonne per ogni strategia
      2. Ontology Jobs     — le 13 professioni O*NET con dettagli
      3. Score Detail      — tutti gli score raw per ogni job × strategia
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = os.path.join(OUTPUT_DIR, "rientra_matching.xlsx")
    wb   = Workbook()

    # Indici per lookup veloce
    s1_idx  = {r["id"]: r for r in s1_res}
    nlp_idx = {key: {r["id"]: r for r in res}
               for key, res in nlp_results.items() if res}
    avail   = [k for k in ["S2_BERT", "S3_BioBERT", "S4_MiniLM"] if k in nlp_idx]

    # ── FOGLIO 1: Matching Results ────────────────────────────
    ws1 = wb.active
    ws1.title = "Matching Results"
    ws1.freeze_panes = "A2"

    # Costruisce header dinamico
    base_cols = [
        (1, "ID"),
        (2, "Job Title (DB)"),
        (3, "Description (DB)"),
        (4, "IRI individuo"),
        # S1
        (5,  "S1 · Match O*NET"),
        (6,  "S1 · Score"),
        (7,  "S1 · Via label"),
        (8,  "S1 · O*NET URI"),
    ]
    nlp_start = 9
    nlp_col_map: dict[str, dict] = {}  # key → {match, score, uri}
    col = nlp_start
    for key in avail:
        short = model_names.get(key, key).split("/")[-1]
        nlp_col_map[key] = {"match": col, "score": col+1, "uri": col+2}
        base_cols += [
            (col,   f"{key[:2]} · {short} Match"),
            (col+1, f"{key[:2]} · Score"),
            (col+2, f"{key[:2]} · O*NET URI"),
        ]
        col += 3

    verdict_col = col
    base_cols.append((verdict_col, "Verdetto"))

    _set_header(ws1, 1, base_cols)

    # Larghezze colonne foglio 1
    col_widths_1 = {1:8, 2:28, 3:55, 4:35, 5:32, 6:10, 7:28, 8:55}
    for key in avail:
        cm = nlp_col_map[key]
        col_widths_1[cm["match"]] = 32
        col_widths_1[cm["score"]] = 10
        col_widths_1[cm["uri"]]   = 55
    col_widths_1[verdict_col] = 12
    for c, w in col_widths_1.items():
        ws1.column_dimensions[get_column_letter(c)].width = w

    for r_idx, row in enumerate(db_jobs, start=2):
        rid  = row["id"]
        alt  = (r_idx % 2 == 0)
        fill = _FILL_ALT if alt else None

        def wc(col, val, cfill=None, cfont=None, wrap=True):
            c = ws1.cell(row=r_idx, column=col, value=val)
            _style_cell(c, fill=cfill or fill, font=cfont, wrap=wrap)

        wc(1, rid,           wrap=False)
        wc(2, row["title"])
        wc(3, row.get("description",""))
        wc(4, f"{RIENTRA_BASE}ExtJob_{rid}")

        # S1
        m1 = s1_idx.get(rid, {}).get("match_s1")
        if m1:
            wc(5, m1["job"]["label"],   cfill=_FILL_S1)
            wc(6, round(m1["score"],4), cfill=_FILL_S1, wrap=False)
            wc(7, m1["via"],            cfill=_FILL_S1)
            wc(8, m1["job"]["uri"],     cfill=_FILL_S1)
        else:
            for c in [5,6,7,8]:
                cell = ws1.cell(row=r_idx, column=c, value="—")
                _style_cell(cell, fill=fill, font=_DIM_FONT, wrap=False)

        # NLP strategies
        uri_set = set()
        for key in avail:
            cm    = nlp_col_map[key]
            r2    = nlp_idx[key].get(rid, {})
            m2    = r2.get("match")
            raw   = r2.get("score_raw", 0.0)
            fills = {"S2_BERT": _FILL_BERT,
                     "S3_BioBERT": _FILL_BIO,
                     "S4_MiniLM": _FILL_MINI}
            sf = fills.get(key)
            if m2:
                wc(cm["match"], m2["job"]["label"],    cfill=sf)
                sc_cell = ws1.cell(row=r_idx, column=cm["score"],
                                   value=round(m2["score"],4))
                _style_cell(sc_cell, fill=sf, wrap=False)
                sc_cell.font = _GREEN_FONT if m2["score"] >= 0.70 else _BODY_FONT
                wc(cm["uri"],   m2["job"]["uri"],       cfill=sf)
                uri_set.add(m2["job"]["uri"])
            else:
                no_cell = ws1.cell(row=r_idx, column=cm["match"],
                                   value=f"sotto soglia (raw {raw:.3f})")
                _style_cell(no_cell, fill=fill, font=_DIM_FONT)
                ws1.cell(row=r_idx, column=cm["score"],
                         value=round(raw,4)).font = _RED_FONT
                _style_cell(ws1.cell(row=r_idx, column=cm["score"]),
                            fill=fill, wrap=False)
                ws1.cell(row=r_idx, column=cm["uri"], value="—")
                _style_cell(ws1.cell(row=r_idx, column=cm["uri"]),
                            fill=fill, font=_DIM_FONT)

        # Verdetto
        if len(uri_set) == 1 and len(uri_set) == len(avail):
            verd, vfont = "Tutti concordano", Font(name="Arial",size=10,
                                                   color="1D6A38", bold=True)
        elif len(uri_set) == 1 and uri_set:
            verd, vfont = "Accordo parziale", Font(name="Arial",size=10,
                                                   color="85610A", bold=True)
        elif len(uri_set) > 1:
            verd, vfont = "Divergono", Font(name="Arial",size=10,
                                            color="A32D2D", bold=True)
        elif not uri_set and m1:
            verd, vfont = "Solo S1", Font(name="Arial",size=10,
                                          color="5C2D91", bold=True)
        else:
            verd, vfont = "Nessun match", Font(name="Arial",size=10,
                                               color="888888", italic=True)
        vc = ws1.cell(row=r_idx, column=verdict_col, value=verd)
        _style_cell(vc, fill=fill, font=vfont, wrap=False)

    ws1.row_dimensions[1].height = 30
    for r in range(2, len(db_jobs)+2):
        ws1.row_dimensions[r].height = 60

    # ── FOGLIO 2: Ontology Jobs ───────────────────────────────
    ws2 = wb.create_sheet("Ontology Jobs")
    ws2.freeze_panes = "A2"
    _set_header(ws2, 1, [
        (1,"#"),(2,"URI locale"),(3,"Label O*NET"),
        (4,"Net job titles (estratto)"),(5,"Descrizione O*NET"),
    ])
    ws2.column_dimensions["A"].width = 4
    ws2.column_dimensions["B"].width = 48
    ws2.column_dimensions["C"].width = 36
    ws2.column_dimensions["D"].width = 50
    ws2.column_dimensions["E"].width = 65

    for i, j in enumerate(onto_jobs, 1):
        alt  = (i % 2 == 0)
        fill = _FILL_ALT if alt else None
        for col, val, wrap in [
            (1, i,               False),
            (2, j["local_name"], True),
            (3, j["label"],      True),
            (4, j["titles"][:300] if j["titles"] else "—", True),
            (5, j["description"][:500] if j["description"] else "—", True),
        ]:
            c = ws2.cell(row=i+1, column=col, value=val)
            _style_cell(c, fill=fill, wrap=wrap)
        ws2.row_dimensions[i+1].height = 60
    ws2.row_dimensions[1].height = 30

    # ── FOGLIO 3: Score Detail ────────────────────────────────
    ws3 = wb.create_sheet("Score Detail")
    ws3.freeze_panes = "C2"

    # Header: ID | Job Title | poi una colonna per ogni modello×professione
    hdr = [(1,"ID"),(2,"Job Title (DB)")]
    col = 3
    model_col_ranges: dict[str, tuple[int,int]] = {}
    for key in avail:
        start = col
        short = model_names.get(key, key).split("/")[-1]
        for j in onto_jobs:
            hdr.append((col, f"{short}\n{j['label'][:22]}"))
            col += 1
        model_col_ranges[key] = (start, col-1)
    _set_header(ws3, 1, hdr)
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 28
    for c in range(3, col):
        ws3.column_dimensions[get_column_letter(c)].width = 14

    # Righe dati con score raw per ogni combinazione
    for r_idx, row in enumerate(db_jobs, start=2):
        rid  = row["id"]
        ws3.cell(row=r_idx, column=1, value=rid)
        ws3.cell(row=r_idx, column=2, value=row["title"])
        for c in [1,2]:
            _style_cell(ws3.cell(row=r_idx, column=c),
                        fill=_FILL_ALT if r_idx%2==0 else None)

        for key in avail:
            r2       = nlp_idx[key].get(rid, {})
            m2       = r2.get("match")
            score_raw = r2.get("score_raw", 0.0)
            start_c, _ = model_col_ranges[key]

            # Per ora salviamo lo score best-match nella colonna della
            # professione trovata, e lasceremo vuote le altre.
            # (score completi richiederebbero di salvare l'intera matrice
            #  nel run_bert_matching — estensione futura)
            best_uri = m2["job"]["uri"] if m2 else None
            for c_off, j in enumerate(onto_jobs):
                c_abs = start_c + c_off
                is_best = (j["uri"] == best_uri)
                val = round(score_raw, 4) if is_best else None
                cell = ws3.cell(row=r_idx, column=c_abs, value=val)
                if is_best and val is not None:
                    fills = {"S2_BERT": _FILL_BERT,
                             "S3_BioBERT": _FILL_BIO,
                             "S4_MiniLM": _FILL_MINI}
                    _style_cell(cell, fill=fills.get(key),
                                font=_GREEN_FONT if score_raw >= 0.70
                                     else _BODY_FONT, wrap=False)
                else:
                    _style_cell(cell,
                                fill=_FILL_ALT if r_idx%2==0 else None,
                                wrap=False)

        ws3.row_dimensions[r_idx].height = 20
    ws3.row_dimensions[1].height = 45

    wb.save(path)
    return path


# ═════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════

def main():
    console.print()
    console.print(Panel(
        f"[bold white]Rientra@[/] — Step 2: RDB2RDF + String Matching + NLP\n"
        f"[dim]Modelli: BERT base · BioBERT · MiniLM  |  Metrica: cosine similarity[/]\n"
        f"[dim]{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}[/]",
        border_style="blue", padding=(0, 2),
    ))
    console.print()

    # ── [0] Setup ─────────────────────────────────────────────
    console.print(Rule("[bold cyan]0 · Setup[/]", style="cyan"))
    console.print()

    if not test_connection():
        sys.exit(1)

    with console.status(f"  Caricamento ontologia [cyan]{ONTOLOGY_FILE}[/]...", spinner="dots"):
        onto_graph = load_ontology(ONTOLOGY_FILE)
    console.print(f"  [bold green]✓[/] Ontologia: [bold]{len(onto_graph)}[/] triple")

    onto_jobs = extract_ontology_jobs(onto_graph)
    console.print(f"  [bold green]✓[/] Professioni O*NET estratte: [bold]{len(onto_jobs)}[/]")
    console.print()
    print_ontology_table(onto_jobs)
    console.print()

    # ── [1] Lettura DB ────────────────────────────────────────
    console.print(Rule("[bold magenta]1 · Lettura DB esterno[/]", style="magenta"))
    console.print()
    db_jobs = fetch_all("SELECT id, title, description FROM ext_job ORDER BY id")
    console.print(f"  [bold green]✓[/] [bold]{len(db_jobs)}[/] lavori da [cyan]ext_job[/]")
    console.print()
    print_db_table(db_jobs)
    console.print()

    # ── [2] Strategia 1 — String Matching ─────────────────────
    console.print(Rule("[bold yellow]2 · Strategia 1 — String Matching[/]", style="yellow"))
    console.print("  [dim]Jaccard (0.5) + Overlap (0.3) + Levenshtein (0.2)[/]")
    console.print()
    s1_res = run_string_matching(db_jobs, onto_jobs)
    console.print(f"  [bold green]✓[/] Completato\n")

    # ── [3] Strategia 2 — BERT base ───────────────────────────
    console.print(Rule("[bold blue]3 · Strategia 2 — BERT base · cosine similarity[/]", style="blue"))
    console.print(f"  [dim]Modello: {MODELS['S2_BERT']}  |  Embedding: token [CLS]  |  Soglia: {BERT_THRESHOLD}[/]")
    console.print()
    s2_res, s2_name = run_bert_matching(
        db_jobs, onto_jobs,
        model_name=MODELS["S2_BERT"],
        strategy_key="S2_BERT",
        threshold=BERT_THRESHOLD,
    )
    if s2_res:
        console.print(f"  [bold green]✓[/] Completato\n")
    else:
        console.print(f"  [yellow]⚠[/] Saltato — installa [cyan]transformers torch[/]\n")

    # ── [4] Strategia 3 — BioBERT ─────────────────────────────
    console.print(Rule("[bold magenta]4 · Strategia 3 — BioBERT · cosine similarity[/]", style="magenta"))
    console.print(f"  [dim]Modello: {MODELS['S3_BioBERT']}  |  Embedding: token [CLS]  |  Soglia: {BERT_THRESHOLD}[/]")
    console.print(f"  [dim]Fine-tuned su: PubMed abstracts + PMC full-text articles (dominio biomedico)[/]")
    console.print()
    s3_res, s3_name = run_bert_matching(
        db_jobs, onto_jobs,
        model_name=MODELS["S3_BioBERT"],
        strategy_key="S3_BioBERT",
        threshold=BERT_THRESHOLD,
    )
    if s3_res:
        console.print(f"  [bold green]✓[/] Completato\n")
    else:
        console.print(f"  [yellow]⚠[/] Saltato\n")

    # ── [5] Strategia 4 — MiniLM ──────────────────────────────
    console.print(Rule("[bold green]5 · Strategia 4 — MiniLM · cosine similarity[/]", style="green"))
    console.print(f"  [dim]Modello: {MODELS['S4_MiniLM']}  |  Fine-tuned per sentence similarity  |  Soglia: {MINILM_THRESHOLD}[/]")
    console.print()
    s4_res, s4_name = run_minilm_matching(db_jobs, onto_jobs, threshold=MINILM_THRESHOLD)
    if s4_res:
        console.print(f"  [bold green]✓[/] Completato\n")
    else:
        console.print(f"  [yellow]⚠[/] Saltato — installa [cyan]sentence-transformers[/]\n")

    # Raccoglie tutti i risultati disponibili
    nlp_results  = {}
    model_names  = {}
    if s2_res: nlp_results["S2_BERT"]    = s2_res;  model_names["S2_BERT"]    = s2_name
    if s3_res: nlp_results["S3_BioBERT"] = s3_res;  model_names["S3_BioBERT"] = s3_name
    if s4_res: nlp_results["S4_MiniLM"]  = s4_res;  model_names["S4_MiniLM"]  = s4_name

    # ── [6] Confronto ─────────────────────────────────────────
    console.print(Rule("[bold white]6 · Confronto strategie[/]", style="white"))
    console.print()
    if nlp_results:
        print_full_comparison(s1_res, nlp_results, model_names)
        print_summary_multi(s1_res, nlp_results, model_names)
    else:
        console.print("  [yellow]Nessun modello NLP disponibile — solo S1 (string matching)[/]")
    console.print()

    # ── [7] Output Excel ──────────────────────────────────────
    console.print(Rule("[bold blue]7 · Produzione file Excel[/]", style="blue"))
    console.print()
    with console.status("  Costruzione workbook...", spinner="dots"):
        xlsx_path = save_excel(db_jobs, s1_res, nlp_results,
                               model_names, onto_jobs)

    size_kb = os.path.getsize(xlsx_path) // 1024
    t = Table(box=box.SIMPLE, show_header=False, padding=(0, 1))
    t.add_column(style="dim")
    t.add_column(style="cyan")
    t.add_column(style="dim", justify="right")
    t.add_row("Excel",          xlsx_path, f"{size_kb:,} KB")
    t.add_row("Foglio 1",       "Matching Results",
              f"{len(db_jobs)} righe · tutte le strategie")
    t.add_row("Foglio 2",       "Ontology Jobs",
              f"{len(onto_jobs)} professioni O*NET")
    t.add_row("Foglio 3",       "Score Detail",
              "score raw per job × modello")
    console.print(t)
    console.print()

    # ── Fine ──────────────────────────────────────────────────
    console.print(Panel(
        f"[bold green]Step 2 completato.[/]\n\n"
        f"Output → [cyan]{xlsx_path}[/]\n\n"
        f"[dim]Modelli usati: {', '.join(model_names.values()) or 'solo S1'}[/]\n"
        f"[dim]Prossimo step: Strategia 3 -- ESCO come ponte ontologico[/]",
        border_style="green", padding=(0, 2),
    ))
    console.print()


if __name__ == "__main__":
    main()